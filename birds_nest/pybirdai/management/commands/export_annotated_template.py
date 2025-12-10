# coding=UTF-8
# Copyright (c) 2025 Bird Software Solutions Ltd
# This program and the accompanying materials
# are made available under the terms of the Eclipse Public License 2.0
# which accompanies this distribution, and is available at
# https://www.eclipse.org/legal/epl-2.0/
#
# SPDX-License-Identifier: EPL-2.0
#
# Contributors:
#    Benjamin Arfa - initial API and implementation

"""
Management command to export an annotated template to Excel.

Usage:
    python manage.py export_annotated_template <table_id>
    python manage.py export_annotated_template <table_id> --output /path/to/output.xlsx
    python manage.py export_annotated_template --list-tables  # List available tables
    python manage.py export_annotated_template --list-tables --framework FINREP
"""

from django.core.management.base import BaseCommand, CommandError
from datetime import datetime
import os
import logging

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = 'Export an annotated template to Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            'table_id',
            nargs='?',
            type=str,
            help='The table ID to export (e.g., FINREP_REF_F_05_01)'
        )
        parser.add_argument(
            '--output', '-o',
            type=str,
            help='Output file path. Defaults to results/annotated_templates/<table_code>_<timestamp>.xlsx'
        )
        parser.add_argument(
            '--list-tables',
            action='store_true',
            help='List available tables instead of exporting'
        )
        parser.add_argument(
            '--framework', '-f',
            type=str,
            help='Filter tables by framework (e.g., FINREP, COREP) when using --list-tables'
        )

    def handle(self, *args, **options):
        if options['list_tables']:
            self.list_tables(options.get('framework'))
            return

        table_id = options['table_id']
        if not table_id:
            raise CommandError('table_id is required. Use --list-tables to see available tables.')

        output_path = options.get('output')
        self.export_template(table_id, output_path)

    def list_tables(self, framework_filter=None):
        """List available tables in the database."""
        from pybirdai.models.bird_meta_data_model import TABLE, FRAMEWORK
        from pybirdai.models.bird_meta_data_model_extension import FRAMEWORK_TABLE

        self.stdout.write(self.style.SUCCESS('Available tables:\n'))

        if framework_filter:
            # Get tables for specific framework
            framework_table_ids = FRAMEWORK_TABLE.objects.filter(
                framework_id__icontains=framework_filter
            ).values_list('table_id', flat=True)
            tables = TABLE.objects.filter(table_id__in=framework_table_ids).order_by('code')
            self.stdout.write(f'Filtered by framework: {framework_filter}\n')
        else:
            tables = TABLE.objects.all().order_by('code')[:100]
            self.stdout.write('Showing first 100 tables. Use --framework to filter.\n')

        self.stdout.write(f'{"Table ID":<60} {"Code":<20} {"Name":<40}')
        self.stdout.write('-' * 120)

        for table in tables:
            self.stdout.write(
                f'{table.table_id:<60} {(table.code or ""):<20} {(table.name or "")[:40]:<40}'
            )

        self.stdout.write(f'\nTotal: {tables.count()} tables')

    def export_template(self, table_id, output_path=None):
        """Export the annotated template to Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise CommandError(
                "openpyxl is required for Excel export. Install with: pip install openpyxl"
            )

        from pybirdai.models.bird_meta_data_model import (
            TABLE, AXIS, AXIS_ORDINATE, ORDINATE_ITEM,
            CELL_POSITION, TABLE_CELL
        )

        # Get the table
        try:
            table = TABLE.objects.get(table_id=table_id)
        except TABLE.DoesNotExist:
            raise CommandError(f'Table not found: {table_id}')

        self.stdout.write(f'Exporting table: {table.code or table_id} - {table.name or ""}')

        # Get axes and ordinates
        table_axes = AXIS.objects.filter(table_id=table)
        table_ordinates = AXIS_ORDINATE.objects.filter(
            axis_id__in=table_axes
        ).select_related('axis_id')

        self.stdout.write(f'  Found {table_axes.count()} axes, {table_ordinates.count()} ordinates')

        # Get ordinate items for annotations
        ordinate_items = ORDINATE_ITEM.objects.filter(
            axis_ordinate_id__in=table_ordinates
        ).select_related('variable_id', 'member_id', 'axis_ordinate_id', 'variable_id__domain_id')

        self.stdout.write(f'  Found {ordinate_items.count()} ordinate items (annotations)')

        # Build ordinate_id -> ordinate items list
        ordinate_items_map = {}
        for item in ordinate_items:
            ord_id = item.axis_ordinate_id.axis_ordinate_id if item.axis_ordinate_id else None
            if ord_id:
                if ord_id not in ordinate_items_map:
                    ordinate_items_map[ord_id] = []
                var_code = item.variable_id.code if item.variable_id else ''
                var_name = item.variable_id.name if item.variable_id else ''
                mem_code = item.member_id.code if item.member_id else ''
                mem_name = item.member_id.name if item.member_id else ''
                domain_code = item.variable_id.domain_id.code if item.variable_id and item.variable_id.domain_id else ''

                ordinate_items_map[ord_id].append({
                    'variable_code': var_code,
                    'variable_name': var_name,
                    'member_code': mem_code,
                    'member_name': mem_name,
                    'domain_code': domain_code,
                })

        # Get cell positions
        cell_positions = CELL_POSITION.objects.filter(
            axis_ordinate_id__in=table_ordinates
        ).select_related('axis_ordinate_id', 'axis_ordinate_id__axis_id', 'cell_id')

        cell_ids = cell_positions.values_list('cell_id', flat=True).distinct()
        table_cells = TABLE_CELL.objects.filter(cell_id__in=cell_ids)

        self.stdout.write(f'  Found {table_cells.count()} cells')

        # Build cell_id -> positions mapping
        cell_to_positions = {}
        for pos in cell_positions:
            cell_id = pos.cell_id_id if hasattr(pos, 'cell_id_id') else (pos.cell_id.cell_id if pos.cell_id else None)
            if cell_id:
                if cell_id not in cell_to_positions:
                    cell_to_positions[cell_id] = []
                cell_to_positions[cell_id].append(pos)

        # Build ordinates data and separate by axis
        ordinates_data = {}
        z_ordinates = []
        row_ordinates_list = []
        col_ordinates_list = []

        for ordinate in table_ordinates:
            orientation = ordinate.axis_id.orientation if ordinate.axis_id else 'Unknown'
            path = ordinate.path or ''
            path_level = path.count('.') if path else 0
            ord_code = ordinate.code or ''
            if not ord_code and '_' in ordinate.axis_ordinate_id:
                parts = ordinate.axis_ordinate_id.split('_')
                ord_code = parts[-1] if parts else ''

            ord_data = {
                'id': ordinate.axis_ordinate_id,
                'name': ordinate.name or ordinate.code or ordinate.axis_ordinate_id,
                'code': ord_code,
                'orientation': orientation,
                'level': path_level,
                'order': ordinate.order or 0,
                'path': path,
            }
            ordinates_data[ordinate.axis_ordinate_id] = ord_data

            if orientation == 'Z':
                z_ordinates.append(ordinate.axis_ordinate_id)
            elif orientation in ['Y', '2']:
                row_ordinates_list.append(ordinate.axis_ordinate_id)
            elif orientation in ['X', '1']:
                col_ordinates_list.append(ordinate.axis_ordinate_id)

        # Build cell matrix
        cell_matrix = {}
        row_ordinates_set = set()
        col_ordinates_set = set()

        for cell in table_cells:
            positions = cell_to_positions.get(cell.cell_id, [])
            row_ord_id = None
            col_ord_id = None

            for pos in positions:
                if pos.axis_ordinate_id and pos.axis_ordinate_id.axis_id:
                    orientation = pos.axis_ordinate_id.axis_id.orientation
                    ord_id = pos.axis_ordinate_id.axis_ordinate_id

                    if orientation in ['Y', '2']:
                        row_ord_id = ord_id
                        row_ordinates_set.add(ord_id)
                    elif orientation in ['X', '1']:
                        col_ord_id = ord_id
                        col_ordinates_set.add(ord_id)

            if row_ord_id and col_ord_id:
                cell_matrix[(row_ord_id, col_ord_id)] = {
                    'cell_id': cell.cell_id,
                    'is_shaded': cell.is_shaded,
                    'name': cell.name or '',
                }

        # Sort ordinates
        row_ordinates = sorted(
            [ord_id for ord_id in row_ordinates_set if ord_id in ordinates_data],
            key=lambda x: (ordinates_data[x].get('level', 0), ordinates_data[x].get('order', 0))
        )
        col_ordinates = sorted(
            [ord_id for ord_id in col_ordinates_set if ord_id in ordinates_data],
            key=lambda x: (ordinates_data[x].get('level', 0), ordinates_data[x].get('order', 0))
        )

        self.stdout.write(f'  Grid: {len(row_ordinates)} rows x {len(col_ordinates)} columns')

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (table.code or 'Template')[:31]  # Excel sheet name limit

        # Styles
        header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
        row_header_fill = PatternFill(start_color='334155', end_color='334155', fill_type='solid')
        shaded_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        dim_header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
        white_font = Font(color='FFFFFF', bold=True)
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Row 1: Title
        ws['A1'] = f"{table.code or table_id} - {table.name or 'Annotated Template'}"
        ws['A1'].font = Font(size=14, bold=True)

        # Row 2+: Z-axis info
        z_row = 2
        if z_ordinates:
            for z_ord_id in z_ordinates:
                z_items = ordinate_items_map.get(z_ord_id, [])
                for z_item in z_items:
                    var_name = z_item['variable_name'] or z_item['variable_code'] or ''
                    var_code = z_item['variable_code'] or ''
                    z_var_info = f"{var_name} ({var_code})" if var_code else var_name
                    ws.cell(row=z_row, column=4, value=z_var_info)
                    ws.cell(row=z_row, column=4).font = bold_font
                    z_row += 1
                    mem_name = z_item['member_name'] or z_item['member_code'] or ''
                    mem_code = z_item['member_code'] or ''
                    z_mem_info = f"{mem_name} ({mem_code})" if mem_code else mem_name
                    ws.cell(row=z_row, column=4, value=z_mem_info)
                    z_row += 1

        # Dynamic positioning
        columns_label_row = max(z_row + 1, 5)
        ws.cell(row=columns_label_row, column=4, value='Columns')
        ws.cell(row=columns_label_row, column=4).font = bold_font

        # Column headers
        header_row = columns_label_row + 1
        for col_idx, col_ord_id in enumerate(col_ordinates, start=4):
            col_data = ordinates_data.get(col_ord_id, {})
            header_text = col_data.get('name', col_ord_id)
            cell = ws.cell(row=header_row, column=col_idx, value=header_text)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = thin_border
            cell.alignment = center_alignment
            col_width = min(max(len(str(header_text)) + 4, 20), 40)
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        # Column codes
        code_row = header_row + 1
        for col_idx, col_ord_id in enumerate(col_ordinates, start=4):
            col_data = ordinates_data.get(col_ord_id, {})
            cell = ws.cell(row=code_row, column=col_idx, value=col_data.get('code', ''))
            cell.border = thin_border
            cell.alignment = center_alignment

        # Collect dimension variables
        z_dim_vars = set()
        for z_ord_id in z_ordinates:
            for item in ordinate_items_map.get(z_ord_id, []):
                if item['variable_code']:
                    z_dim_vars.add(item['variable_code'])

        y_dim_vars = set()
        for row_ord_id in row_ordinates:
            for item in ordinate_items_map.get(row_ord_id, []):
                if item['variable_code'] and item['variable_code'] not in z_dim_vars:
                    y_dim_vars.add(item['variable_code'])
        y_dim_vars_list = sorted(y_dim_vars)

        y_var_names = {}
        for row_ord_id in row_ordinates:
            for item in ordinate_items_map.get(row_ord_id, []):
                if item['variable_code'] and item['variable_code'] not in y_var_names:
                    y_var_names[item['variable_code']] = item['variable_name'] or item['variable_code']

        # Y-axis dimension headers
        dim_start_col = 4 + len(col_ordinates) + 2
        for dim_idx, dim_var in enumerate(y_dim_vars_list):
            var_name = y_var_names.get(dim_var, dim_var)
            header_text = f"{var_name} ({dim_var})" if var_name != dim_var else dim_var
            cell = ws.cell(row=header_row, column=dim_start_col + dim_idx, value=header_text)
            cell.fill = dim_header_fill
            cell.font = white_font
            cell.border = thin_border
            cell.alignment = center_alignment
            dim_col_width = min(max(len(str(header_text)) + 4, 35), 50)
            ws.column_dimensions[get_column_letter(dim_start_col + dim_idx)].width = dim_col_width

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 45

        ws.row_dimensions[header_row].height = 40
        ws.row_dimensions[code_row].height = 20
        ws.freeze_panes = ws.cell(row=code_row + 1, column=4)

        # Data rows
        data_start_row = code_row + 1
        for row_idx, row_ord_id in enumerate(row_ordinates, start=data_start_row):
            row_data = ordinates_data.get(row_ord_id, {})

            if row_idx == data_start_row:
                ws.cell(row=row_idx, column=1, value='Rows')
                ws.cell(row=row_idx, column=1).font = bold_font

            level = row_data.get('level', 0)
            indent = '  ' * level
            header_text = row_data.get('name', row_ord_id)
            cell = ws.cell(row=row_idx, column=2, value=f"{indent}{header_text}")
            cell.fill = row_header_fill
            cell.font = white_font
            cell.border = thin_border
            cell.alignment = left_alignment

            cell = ws.cell(row=row_idx, column=3, value=row_data.get('code', ''))
            cell.border = thin_border
            cell.alignment = center_alignment

            for col_idx, col_ord_id in enumerate(col_ordinates, start=4):
                cell_key = (row_ord_id, col_ord_id)
                cell_info = cell_matrix.get(cell_key)

                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_alignment

                if cell_info:
                    cell_id = cell_info.get('cell_id', '')
                    if cell_id.startswith('REF_'):
                        display_id = cell_id.replace('REF_', '')
                    else:
                        display_id = cell_id
                    cell.value = f"{display_id}\n$"
                    if cell_info.get('is_shaded'):
                        cell.fill = shaded_fill
                else:
                    cell.value = ''

            # Y-axis dimension columns
            row_items = ordinate_items_map.get(row_ord_id, [])
            row_dim_values = {item['variable_code']: item for item in row_items if item['variable_code']}

            for dim_idx, dim_var in enumerate(y_dim_vars_list):
                if dim_var in row_dim_values:
                    item = row_dim_values[dim_var]
                    mem_name = item['member_name'] or item['member_code'] or '*'
                    mem_code = item['member_code'] or ''
                    dim_text = f"{mem_name} ({mem_code})" if mem_code else mem_name
                    cell = ws.cell(row=row_idx, column=dim_start_col + dim_idx, value=dim_text)
                    cell.border = thin_border
                    cell.alignment = left_alignment

        # X-axis ordinate item rows
        x_dim_vars = set()
        for col_ord_id in col_ordinates:
            for item in ordinate_items_map.get(col_ord_id, []):
                if item['variable_code'] and item['variable_code'] not in z_dim_vars:
                    x_dim_vars.add(item['variable_code'])
        x_dim_vars_list = sorted(x_dim_vars)

        x_dim_start_row = data_start_row + len(row_ordinates) + 1
        x_var_names = {}
        for col_ord_id in col_ordinates:
            for item in ordinate_items_map.get(col_ord_id, []):
                if item['variable_code'] and item['variable_code'] not in x_var_names:
                    x_var_names[item['variable_code']] = item['variable_name'] or item['variable_code']

        for x_dim_idx, x_dim_var in enumerate(x_dim_vars_list):
            current_row = x_dim_start_row + x_dim_idx
            var_name = x_var_names.get(x_dim_var, x_dim_var)
            label_text = f"{var_name} ({x_dim_var})" if var_name != x_dim_var else x_dim_var
            label_cell = ws.cell(row=current_row, column=2, value=label_text)
            label_cell.font = white_font
            label_cell.fill = dim_header_fill
            label_cell.border = thin_border
            label_cell.alignment = left_alignment

            for col_idx, col_ord_id in enumerate(col_ordinates, start=4):
                col_items = ordinate_items_map.get(col_ord_id, [])
                item_for_var = None
                for item in col_items:
                    if item['variable_code'] == x_dim_var:
                        item_for_var = item
                        break

                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = thin_border
                cell.alignment = left_alignment
                if item_for_var:
                    mem_name = item_for_var['member_name'] or item_for_var['member_code'] or '*'
                    mem_code = item_for_var['member_code'] or ''
                    cell.value = f"{mem_name} ({mem_code})" if mem_code else mem_name

        # Determine output path
        if not output_path:
            output_dir = os.path.join('results', 'annotated_templates')
            os.makedirs(output_dir, exist_ok=True)
            safe_table_code = (table.code or table_id).replace('/', '_').replace('\\', '_').replace('.', '_')
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = os.path.join(output_dir, f"annotated_template_{safe_table_code}_{timestamp}.xlsx")

        # Ensure parent directory exists
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

        # Save workbook
        wb.save(output_path)

        self.stdout.write(self.style.SUCCESS(f'\nExported to: {output_path}'))
