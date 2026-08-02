import csv
import datetime
import io
import json
import os
import re
import types
import unittest
import zipfile
from types import SimpleNamespace
from unittest.mock import patch
from xml.etree import ElementTree

from django.db import models
from django.test import RequestFactory, SimpleTestCase
from openpyxl import Workbook, load_workbook

from pybirdai.utils.datapoint_test_run.excel_to_csv_converter import (
    MAX_WORKBOOK_BYTES,
    ExcelToCSVConverter,
    UnsafeCellValueError,
    WorkbookTooLargeError,
)
from pybirdai.utils.datapoint_test_run.test_data_template_utils import (
    DROPDOWN_MAX_LENGTH,
    build_domain_code_lookup,
    format_domain_value_for_cell,
    format_domain_value_for_dropdown,
    format_value_for_excel,
    get_local_field_export_map,
    get_model_fields_metadata,
    resolve_domain_value_to_code,
)
from pybirdai.views.test_data_template_views import (
    EXCEL_MAX_SHEET_TITLE_LENGTH,
    _make_unique_excel_sheet_title,
    _read_existing_rows,
    export_bird_excel_template,
    import_excel_test_data,
)

STATUS_DOMAIN = {
    "14": "Not_in_default",
    "18": "Default_because_both_unlikely_to_pay_and_past_due",
    "20": "Ratio: more than 90 days past due",
}


def _build_field(field, name):
    """Attach the naming Django would give a field bound to a model."""
    field.name = name
    field.attname = name
    field.column = name
    return field


class _FakeModelClass:
    """
    Attribute bag standing in for a generated model class.

    Unlike SimpleNamespace it stays hashable, so it can be used as a key the way
    Django uses model classes in _meta.parents.
    """

    def __init__(self, **attributes):
        self.__dict__.update(attributes)


def _fake_model(name, local_fields, inherited_fields=(), domains=None, parents=None):
    """
    Build a stand-in for a generated BIRD model.

    The converter only reads field metadata, never the database, so real Django
    field objects plus a minimal _meta are enough - and this keeps the tests
    independent of whether bird_data_model.py has been generated.

    `parents` mirrors Django's _meta.parents: {parent model: parent link field}.
    """
    all_fields = list(local_fields) + list(inherited_fields)
    model = _FakeModelClass(
        __name__=name,
        _meta=types.SimpleNamespace(
            local_fields=list(local_fields),
            get_fields=lambda: all_fields,
            db_table=f"pybirdai_{name.lower()}",
            parents=dict(parents or {}),
        ),
    )
    for domain_name, domain in (domains or {}).items():
        setattr(model, domain_name, domain)
    return model


class ExcelSheetTitleTests(unittest.TestCase):
    def test_sheet_titles_remain_unique_after_truncation(self):
        used_titles = set()

        first = _make_unique_excel_sheet_title("a" * 40, used_titles)
        second = _make_unique_excel_sheet_title("A" * 40, used_titles)
        third = _make_unique_excel_sheet_title("bad/name:*?[]\\'title", used_titles)

        self.assertEqual(len(first), EXCEL_MAX_SHEET_TITLE_LENGTH)
        self.assertLessEqual(len(second), EXCEL_MAX_SHEET_TITLE_LENGTH)
        self.assertNotEqual(first.casefold(), second.casefold())
        self.assertNotRegex(third, r"[\\/*?:\[\]]")


class ExcelTestDataTemplateTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()

    @patch("pybirdai.utils.datapoint_test_run.test_data_template_utils.extract_domain_dictionaries")
    @patch("pybirdai.utils.datapoint_test_run.test_data_template_utils.get_model_fields_metadata")
    @patch("pybirdai.utils.datapoint_test_run.test_data_template_utils.get_bird_model_classes")
    def test_export_uses_excel_safe_titles_and_range_backed_validations(
        self,
        get_models,
        get_fields,
        get_domains,
    ):
        model_names = [
            "LONG_MODEL_NAME_WITH_A_COLLIDING_PREFIX_ALPHA",
            "LONG_MODEL_NAME_WITH_A_COLLIDING_PREFIX_BETA",
        ]
        get_models.return_value = {name: object() for name in model_names}
        get_fields.return_value = [
            {
                "name": "STATUS",
                "python_type": "str",
                "is_foreign_key": False,
                "related_model": None,
            }
        ]
        get_domains.return_value = {
            "STATUS": {str(code): f'Description_with_a_comma,_quote_"_and_value_{code}' for code in range(1, 25)}
        }

        request = self.factory.get(
            "/api/test-data/excel-template/",
            {"include_all": "true"},
        )
        response = export_bird_excel_template(request)

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content))

        self.assertEqual(
            len(workbook.sheetnames),
            len({title.casefold() for title in workbook.sheetnames}),
        )
        for title in workbook.sheetnames:
            self.assertLessEqual(len(title), EXCEL_MAX_SHEET_TITLE_LENGTH)
            self.assertIsNone(re.search(r"[\\/*?:\[\]]", title))

        table_index = workbook["_Table_Index"]
        indexed_rows = list(
            table_index.iter_rows(
                min_row=2,
                max_row=1 + len(model_names),
                values_only=True,
            )
        )
        self.assertEqual([row[0] for row in indexed_rows], model_names)
        self.assertNotEqual(indexed_rows[0][1], indexed_rows[1][1])
        self.assertTrue(all(len(row[1]) <= EXCEL_MAX_SHEET_TITLE_LENGTH for row in indexed_rows))

        validation_sheet = workbook["_Validation_Lists"]
        self.assertEqual(validation_sheet.sheet_state, "veryHidden")
        self.assertIn("BIRD_Domain_1", workbook.defined_names)
        self.assertIn(
            "'_Validation_Lists'!",
            workbook.defined_names["BIRD_Domain_1"].attr_text,
        )

        for _, sheet_name, _included_as in indexed_rows:
            validations = list(workbook[sheet_name].data_validations.dataValidation)
            self.assertEqual(len(validations), 1)
            self.assertEqual(validations[0].formula1, "=BIRD_Domain_1")

        # Every XML relationship and content part should remain well-formed.
        with zipfile.ZipFile(io.BytesIO(response.content)) as package:
            xml_parts = [name for name in package.namelist() if name.endswith((".xml", ".rels"))]
            self.assertTrue(xml_parts)
            for part_name in xml_parts:
                ElementTree.fromstring(package.read(part_name))


def _parent_link_field(name, parent_name):
    """A stand-in for the link Django adds to a multi-table-inheritance child."""
    field = _build_field(
        models.OneToOneField(parent_name, on_delete=models.CASCADE, parent_link=True, primary_key=True),
        name,
    )
    # An unbound field's related_model is still the model name as text.
    field.related_model = types.SimpleNamespace(__name__=parent_name)
    return field


class LocalColumnsOnlyTests(SimpleTestCase):
    """A worksheet offers the same columns the fixture CSV for that table holds."""

    def setUp(self):
        self.factory = RequestFactory()

    def test_metadata_describes_a_tables_own_columns_only(self):
        model = _fake_model(
            "CHILD",
            local_fields=[
                _parent_link_field("parent_ptr", "PARENT"),
                _build_field(models.CharField(max_length=255), "OWN"),
            ],
            inherited_fields=[_build_field(models.CharField(max_length=255), "INHERITED")],
        )

        headers = [meta["name"] for meta in get_model_fields_metadata(model)]

        self.assertEqual(headers, ["parent_ptr", "OWN"])
        self.assertNotIn("INHERITED", headers)

        # Every header the worksheet offers has to be one the converter stores,
        # otherwise the value entered under it is silently dropped.
        export_columns = get_local_field_export_map(model)
        self.assertEqual(len(headers), len(export_columns))
        for header, export_column in zip(headers, export_columns):
            self.assertIn(header, export_column["accepted_names"])

    def test_export_adds_the_parent_table_of_a_selected_subtype(self):
        parent = _fake_model(
            "PARENT",
            local_fields=[_build_field(models.CharField(max_length=255), "PARENT_uniqueID")],
        )
        parent_link = _parent_link_field("parent_ptr", "PARENT")
        child = _fake_model(
            "CHILD",
            local_fields=[parent_link, _build_field(models.CharField(max_length=255), "OWN")],
            inherited_fields=[_build_field(models.CharField(max_length=255), "PARENT_uniqueID")],
            parents={parent: parent_link},
        )

        with patch(
            "pybirdai.utils.datapoint_test_run.test_data_template_utils.get_bird_model_classes",
            return_value={"PARENT": parent, "CHILD": child},
        ):
            request = self.factory.get(
                "/api/test-data/excel-template/",
                {"tables": "CHILD", "include_data": "false"},
            )
            response = export_bird_excel_template(request)

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content))

        # The subtype was asked for; its parent came along so that the fields
        # declared on the parent still have a worksheet to be entered on.
        self.assertIn("child", workbook.sheetnames)
        self.assertIn("parent", workbook.sheetnames)

        self.assertEqual(
            [cell.value for cell in workbook["child"][1]],
            ["parent_ptr", "OWN"],
        )
        self.assertEqual(
            [cell.value for cell in workbook["parent"][1]],
            ["PARENT_uniqueID"],
        )

        index_rows = list(workbook["_Table_Index"].iter_rows(min_row=2, values_only=True))
        self.assertIn(("CHILD", "child", "selected"), index_rows)
        self.assertIn(("PARENT", "parent", "parent table"), index_rows)


class DomainCodeRoundTripTests(unittest.TestCase):
    """The dropdown shows "code: description"; the database stores the code."""

    def setUp(self):
        self.lookup = build_domain_code_lookup(STATUS_DOMAIN)

    def test_dropdown_label_resolves_back_to_its_code(self):
        for code, description in STATUS_DOMAIN.items():
            label = format_domain_value_for_dropdown(code, description, DROPDOWN_MAX_LENGTH)
            self.assertEqual(resolve_domain_value_to_code(label, self.lookup), code)

    def test_a_bare_code_is_accepted(self):
        self.assertEqual(resolve_domain_value_to_code("14", self.lookup), "14")
        self.assertEqual(resolve_domain_value_to_code("  14  ", self.lookup), "14")

    def test_a_description_is_accepted_with_underscores_or_spaces(self):
        self.assertEqual(resolve_domain_value_to_code("Not_in_default", self.lookup), "14")
        self.assertEqual(resolve_domain_value_to_code("Not in default", self.lookup), "14")

    def test_a_description_containing_a_colon_still_resolves(self):
        # The code always comes first, so splitting on the first colon is safe
        # even when the description has one of its own.
        label = format_domain_value_for_dropdown("20", STATUS_DOMAIN["20"], DROPDOWN_MAX_LENGTH)
        self.assertIn(":", STATUS_DOMAIN["20"])
        self.assertEqual(resolve_domain_value_to_code(label, self.lookup), "20")

    def test_a_truncated_label_still_resolves(self):
        long_domain = {"7": "A_" + ("very_long_" * 40) + "description"}
        lookup = build_domain_code_lookup(long_domain)
        label = format_domain_value_for_dropdown("7", long_domain["7"], DROPDOWN_MAX_LENGTH)

        self.assertTrue(label.endswith("..."))
        self.assertEqual(resolve_domain_value_to_code(label, lookup), "7")

    def test_labels_that_truncate_to_the_same_text_stay_distinct(self):
        shared = "Shared_" + ("prefix_" * 40)
        domain = {"1": shared + "one", "2": shared + "two"}
        lookup = build_domain_code_lookup(domain)

        first = format_domain_value_for_dropdown("1", domain["1"], DROPDOWN_MAX_LENGTH)
        second = format_domain_value_for_dropdown("2", domain["2"], DROPDOWN_MAX_LENGTH)

        self.assertNotEqual(first, second)
        self.assertEqual(resolve_domain_value_to_code(first, lookup), "1")
        self.assertEqual(resolve_domain_value_to_code(second, lookup), "2")

    def test_unrecognised_and_empty_values_return_none(self):
        self.assertIsNone(resolve_domain_value_to_code("999: nonsense", self.lookup))
        self.assertIsNone(resolve_domain_value_to_code("", self.lookup))
        self.assertIsNone(resolve_domain_value_to_code(None, self.lookup))

    def test_a_stored_code_is_written_as_the_dropdown_label(self):
        self.assertEqual(
            format_domain_value_for_cell("14", STATUS_DOMAIN),
            format_domain_value_for_dropdown("14", "Not_in_default", DROPDOWN_MAX_LENGTH),
        )

    def test_a_code_outside_the_domain_is_written_unchanged(self):
        # Pre-existing data can hold a value the domain does not list; it must
        # survive the round trip rather than be dropped.
        self.assertEqual(format_domain_value_for_cell("64_1", STATUS_DOMAIN), "64_1")
        self.assertEqual(format_domain_value_for_cell(None, STATUS_DOMAIN), "")


class ExistingDataExportTests(SimpleTestCase):
    def test_rows_are_written_with_dropdown_labels_and_raw_foreign_keys(self):
        fields_metadata = [
            {"name": "STATUS", "is_foreign_key": False},
            {"name": "AMOUNT", "is_foreign_key": False},
            {"name": "WHEN", "is_foreign_key": False},
            {"name": "theOTHER", "is_foreign_key": True},
        ]
        instance = types.SimpleNamespace(
            STATUS="14",
            AMOUNT=1200,
            WHEN=datetime.datetime(2018, 9, 30, 0, 0, 0),
            theOTHER_id="OTHER_1",
        )
        model_class = types.SimpleNamespace(
            __name__="THING",
            objects=types.SimpleNamespace(all=lambda: [instance]),
        )

        rows = _read_existing_rows(model_class, fields_metadata, {"STATUS": STATUS_DOMAIN}, 10)

        self.assertEqual(
            rows,
            [
                [
                    format_domain_value_for_dropdown("14", "Not_in_default", DROPDOWN_MAX_LENGTH),
                    1200,
                    "2018-09-30 00:00:00",
                    "OTHER_1",
                ]
            ],
        )

    def test_a_table_that_cannot_be_read_contributes_no_rows(self):
        class Unreadable:
            __name__ = "UNREADABLE"

            class objects:
                @staticmethod
                def all():
                    raise RuntimeError("table does not exist")

        self.assertEqual(_read_existing_rows(Unreadable, [{"name": "A"}], {}, 10), [])

    def test_export_can_be_asked_for_an_empty_template(self):
        with patch(
            "pybirdai.views.test_data_template_views._read_existing_rows"
        ) as read_rows:
            with patch(
                "pybirdai.utils.datapoint_test_run.test_data_template_utils.get_bird_model_classes",
                return_value={"THING": object()},
            ), patch(
                "pybirdai.utils.datapoint_test_run.test_data_template_utils.get_model_fields_metadata",
                return_value=[{"name": "STATUS", "is_foreign_key": False}],
            ), patch(
                "pybirdai.utils.datapoint_test_run.test_data_template_utils.extract_domain_dictionaries",
                return_value={},
            ):
                request = RequestFactory().get(
                    "/api/test-data/excel-template/",
                    {"include_all": "true", "include_data": "false"},
                )
                response = export_bird_excel_template(request)

        self.assertEqual(response.status_code, 200)
        read_rows.assert_not_called()


class ExcelToCSVConverterTests(SimpleTestCase):
    def _thing_model(self):
        return _fake_model(
            "THING",
            local_fields=[
                _build_field(models.CharField(max_length=255), "THING_uniqueID"),
                _build_field(models.CharField(max_length=255, choices=STATUS_DOMAIN), "STATUS"),
                _build_field(models.BigIntegerField(), "AMOUNT"),
                _build_field(models.FloatField(), "RATE"),
                _build_field(models.DateTimeField(), "WHEN"),
                _build_field(models.BooleanField(), "FLAG"),
            ],
            inherited_fields=[_build_field(models.CharField(max_length=255), "INHERITED")],
            domains={"STATUS_domain": STATUS_DOMAIN},
        )

    def _workbook(self, rows, headers=None, sheet_title="thing"):
        workbook = Workbook()
        workbook.remove(workbook.active)
        sheet = workbook.create_sheet(title=sheet_title)
        sheet.append(
            headers or ["THING_uniqueID", "STATUS", "AMOUNT", "RATE", "WHEN", "FLAG", "INHERITED"]
        )
        for row in rows:
            sheet.append(row)

        index = workbook.create_sheet(title="_Table_Index", index=0)
        index.append(["BIRD model/table", "Worksheet"])
        index.append(["THING", sheet_title])

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def _convert(self, workbook_bytes, model=None):
        model = model or self._thing_model()
        with patch(
            "pybirdai.utils.datapoint_test_run.excel_to_csv_converter.get_bird_model_classes",
            return_value={"THING": model},
        ), patch(
            "pybirdai.utils.datapoint_test_run.excel_to_csv_converter.get_model_by_table_name",
            return_value=model,
        ):
            return ExcelToCSVConverter().convert_workbook(workbook_bytes)

    def test_dropdown_labels_become_codes_and_inherited_columns_are_dropped(self):
        label = format_domain_value_for_dropdown("18", STATUS_DOMAIN["18"], DROPDOWN_MAX_LENGTH)
        workbook = self._workbook(
            [
                ["ID_1", label, 1200, 2.0, datetime.datetime(2018, 9, 30), True, "from parent"],
                ["ID_2", "14", 0, 0.5, datetime.date(2019, 1, 31), False, "from parent"],
            ]
        )

        result = self._convert(workbook)

        rows = list(csv.DictReader(io.StringIO(result["files"]["thing.csv"])))
        self.assertEqual([row["STATUS"] for row in rows], ["18", "14"])
        self.assertNotIn("INHERITED", rows[0])
        self.assertEqual(result["ignored_columns"], {"THING": ["INHERITED"]})
        self.assertEqual(result["tables"], {"THING": 2})
        self.assertEqual(result["unresolved_values"], [])

    def test_numbers_dates_and_booleans_keep_the_fixture_format(self):
        workbook = self._workbook(
            [["ID_1", "14", 1200, 2.0, datetime.datetime(2018, 9, 30, 12, 30), True, ""]]
        )

        result = self._convert(workbook)
        row = next(csv.DictReader(io.StringIO(result["files"]["thing.csv"])))

        # A whole number must not gain a ".0", but a float field must keep one.
        self.assertEqual(row["AMOUNT"], "1200")
        self.assertEqual(row["RATE"], "2.0")
        self.assertEqual(row["WHEN"], "2018-09-30 12:30:00")
        self.assertEqual(row["FLAG"], "True")

    def test_blank_rows_and_null_markers_become_empty_cells(self):
        workbook = self._workbook(
            [
                ["ID_1", "", "NULL", None, None, None, None],
                [None, None, None, None, None, None, None],
                ["   ", "", "", "", "", "", ""],
            ]
        )

        result = self._convert(workbook)
        rows = list(csv.DictReader(io.StringIO(result["files"]["thing.csv"])))

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["STATUS"], "")
        self.assertEqual(rows[0]["AMOUNT"], "")

    def test_an_unrecognised_dropdown_value_is_kept_and_reported(self):
        workbook = self._workbook([["ID_1", "64_1", 1, 1.0, None, None, None]])

        result = self._convert(workbook)
        row = next(csv.DictReader(io.StringIO(result["files"]["thing.csv"])))

        self.assertEqual(row["STATUS"], "64_1")
        self.assertEqual(len(result["unresolved_values"]), 1)
        self.assertIn("64_1", result["unresolved_values"][0])

    def test_guidance_sheets_are_ignored_and_unknown_sheets_reported(self):
        workbook = Workbook()
        workbook.remove(workbook.active)
        workbook.create_sheet(title="_Instructions").append(["BIRD Test Data Template"])
        workbook.create_sheet(title="not_a_bird_table").append(["A"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        with patch(
            "pybirdai.utils.datapoint_test_run.excel_to_csv_converter.get_bird_model_classes",
            return_value={},
        ), patch(
            "pybirdai.utils.datapoint_test_run.excel_to_csv_converter.get_model_by_table_name",
            return_value=None,
        ):
            result = ExcelToCSVConverter().convert_workbook(buffer)

        self.assertEqual(result["skipped_sheets"], ["not_a_bird_table"])
        self.assertEqual(result["files"], {})


class ImportExcelTestDataViewTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def test_a_workbook_is_returned_as_a_zip_of_csv_fixtures(self):
        converter_result = {
            "files": {"thing.csv": "THING_uniqueID,STATUS\nID_1,14\n"},
            "tables": {"THING": 1},
            "skipped_sheets": [],
            "ignored_columns": {},
            "unresolved_values": [],
        }
        upload = io.BytesIO(b"not really a workbook")
        upload.name = "test_data.xlsx"

        with patch(
            "pybirdai.utils.datapoint_test_run.excel_to_csv_converter.ExcelToCSVConverter.convert_workbook",
            return_value=converter_result,
        ):
            request = self.factory.post("/api/test-data/import-excel/", {"file": upload})
            response = import_excel_test_data(request)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")

        with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
            self.assertIn("thing.csv", archive.namelist())
            self.assertIn("_conversion_report.json", archive.namelist())
            self.assertEqual(archive.read("thing.csv").decode(), converter_result["files"]["thing.csv"])

    def test_a_missing_or_wrongly_typed_upload_is_rejected(self):
        response = import_excel_test_data(self.factory.post("/api/test-data/import-excel/"))
        self.assertEqual(response.status_code, 400)

        upload = io.BytesIO(b"a,b\n1,2\n")
        upload.name = "data.csv"
        response = import_excel_test_data(
            self.factory.post("/api/test-data/import-excel/", {"file": upload})
        )
        self.assertEqual(response.status_code, 400)

    def test_an_unreadable_workbook_reports_a_clean_error(self):
        upload = io.BytesIO(b"not really a workbook")
        upload.name = "broken.xlsx"

        response = import_excel_test_data(
            self.factory.post("/api/test-data/import-excel/", {"file": upload})
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("could not be read", response.content.decode())


class ExcelValueFormattingTests(unittest.TestCase):
    def test_values_are_written_in_the_format_the_fixture_loader_expects(self):
        self.assertEqual(format_value_for_excel(None), "")
        self.assertEqual(format_value_for_excel(True), "True")
        self.assertEqual(format_value_for_excel(datetime.date(2018, 9, 30)), "2018-09-30")
        self.assertEqual(
            format_value_for_excel(datetime.datetime(2018, 9, 30, 1, 2, 3)),
            "2018-09-30 01:02:03",
        )
        # Numbers stay numbers so Excel does not left-align them as text.
        self.assertEqual(format_value_for_excel(1200), 1200)
        self.assertEqual(format_value_for_excel(2.5), 2.5)


class WorkbookSecurityTests(SimpleTestCase):
    """The workbook is untrusted input whose output reaches a database."""

    def _thing_model(self):
        return _fake_model(
            "THING",
            local_fields=[
                _build_field(models.CharField(max_length=255), "THING_uniqueID"),
                _build_field(models.CharField(max_length=255, choices=STATUS_DOMAIN), "STATUS"),
            ],
            domains={"STATUS_domain": STATUS_DOMAIN},
        )

    def _workbook_with(self, value, header="THING_uniqueID", as_text=False):
        workbook = Workbook()
        workbook.remove(workbook.active)
        sheet = workbook.create_sheet(title="thing")
        sheet.append([header, "STATUS"])
        sheet.append([None, "14"])
        cell = sheet.cell(row=2, column=1)
        if as_text:
            # A cell openpyxl would otherwise store as a formula. Forced to
            # text it survives the round trip, which is the hostile case: the
            # spreadsheet only evaluates it when the CSV is opened later.
            cell._value = value
            cell.data_type = "s"
        else:
            cell.value = value
        index = workbook.create_sheet(title="_Table_Index", index=0)
        index.append(["BIRD model/table", "Worksheet"])
        index.append(["THING", "thing"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def _convert(self, workbook_bytes):
        model = self._thing_model()
        with patch(
            "pybirdai.utils.datapoint_test_run.excel_to_csv_converter.get_bird_model_classes",
            return_value={"THING": model},
        ), patch(
            "pybirdai.utils.datapoint_test_run.excel_to_csv_converter.get_model_by_table_name",
            return_value=model,
        ):
            return ExcelToCSVConverter().convert_workbook(workbook_bytes)

    def test_a_formula_cell_is_refused_rather_than_written(self):
        # Writing this into a CSV would run it if the CSV were opened in a
        # spreadsheet, so the conversion stops and names the cell.
        for formula in ("=cmd|'/c calc'!A0", "+1+1", "@SUM(A1:A9)"):
            with self.subTest(formula=formula):
                with self.assertRaises(UnsafeCellValueError) as caught:
                    self._convert(self._workbook_with(formula, as_text=True))
                self.assertIn("thing!THING_uniqueID row 2", str(caught.exception))

    def test_a_real_formula_never_reaches_the_csv_as_its_text(self):
        # A cell openpyxl stores as a formula is read for its cached result, so
        # the formula text cannot be written out even when it is not rejected.
        result = self._convert(self._workbook_with("=1+1"))

        row = next(csv.DictReader(io.StringIO(result["files"]["thing.csv"])))
        self.assertNotIn("=", row["THING_uniqueID"])

    def test_a_cell_holding_a_line_break_is_refused(self):
        with self.assertRaises(UnsafeCellValueError):
            self._convert(self._workbook_with("ID_1\r\nID_2"))

    def test_ordinary_values_including_negative_numbers_are_untouched(self):
        # A spreadsheet returns a negative number as a number, so the formula
        # check never sees a leading '-'.
        result = self._convert(self._workbook_with(-42))

        row = next(csv.DictReader(io.StringIO(result["files"]["thing.csv"])))
        self.assertEqual(row["THING_uniqueID"], "-42")

    def test_an_oversized_workbook_is_refused_before_it_is_parsed(self):
        oversized = io.BytesIO(b"x")
        oversized.size = MAX_WORKBOOK_BYTES + 1

        with self.assertRaises(WorkbookTooLargeError):
            ExcelToCSVConverter().convert_workbook(oversized)

    def test_too_many_worksheets_are_refused(self):
        workbook = Workbook()
        workbook.remove(workbook.active)
        for index in range(3):
            workbook.create_sheet(title=f"sheet_{index}")
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        with patch(
            "pybirdai.utils.datapoint_test_run.excel_to_csv_converter.MAX_WORKSHEETS", 2
        ):
            with self.assertRaises(WorkbookTooLargeError):
                ExcelToCSVConverter().convert_workbook(buffer)

    def test_too_many_rows_are_refused(self):
        workbook = Workbook()
        workbook.remove(workbook.active)
        sheet = workbook.create_sheet(title="thing")
        sheet.append(["THING_uniqueID", "STATUS"])
        for index in range(5):
            sheet.append([f"ID_{index}", "14"])
        index_sheet = workbook.create_sheet(title="_Table_Index", index=0)
        index_sheet.append(["BIRD model/table", "Worksheet"])
        index_sheet.append(["THING", "thing"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        with patch(
            "pybirdai.utils.datapoint_test_run.excel_to_csv_converter.MAX_ROWS_PER_SHEET", 2
        ):
            with self.assertRaises(WorkbookTooLargeError):
                self._convert(buffer)

    def test_a_column_the_model_does_not_have_cannot_reach_the_csv(self):
        # The column set comes from the model, so an added header is reported
        # and dropped rather than becoming a new database column.
        result = self._convert(self._workbook_with("ID_1", header="DROP TABLE pybirdai_prty"))

        self.assertEqual(result["ignored_columns"], {"THING": ["DROP TABLE pybirdai_prty"]})
        columns = next(csv.reader(io.StringIO(result["files"]["thing.csv"])))
        self.assertEqual(columns, ["THING_uniqueID", "STATUS"])


class ImportExcelTargetTests(SimpleTestCase):
    """The CSVs may only be written into a fixture scenario that already exists."""

    def setUp(self):
        self.factory = RequestFactory()
        self.converter_result = {
            "files": {"thing.csv": "THING_uniqueID,STATUS\nID_1,14\n"},
            "tables": {"THING": 1},
            "skipped_sheets": [],
            "ignored_columns": {},
            "unresolved_values": [],
        }

    def _upload(self):
        upload = io.BytesIO(b"not really a workbook")
        upload.name = "test_data.xlsx"
        return upload

    def test_an_unknown_scenario_key_is_refused(self):
        with patch(
            "pybirdai.views.test_data_template_views.find_fixture_scenario", return_value=None
        ), patch(
            "pybirdai.utils.datapoint_test_run.excel_to_csv_converter.ExcelToCSVConverter.convert_workbook"
        ) as convert:
            request = self.factory.post(
                "/api/test-data/import-excel/",
                {"file": self._upload(), "scenario_key": "made/up"},
            )
            response = import_excel_test_data(request)

        self.assertEqual(response.status_code, 404)
        # The workbook is not even parsed when the target is unusable.
        convert.assert_not_called()

    def test_a_known_scenario_is_written_to_by_name_not_by_path(self):
        scenario = SimpleNamespace(
            key="suite-a/TEMPLATE/scenario_1",
            name="scenario_1",
            project_path=os.path.join("tests", "suite-a", "scenario_1"),
        )

        with patch(
            "pybirdai.views.test_data_template_views.find_fixture_scenario", return_value=scenario
        ), patch(
            "pybirdai.utils.datapoint_test_run.excel_to_csv_converter.ExcelToCSVConverter.convert_workbook",
            return_value=self.converter_result,
        ), patch(
            "pybirdai.views.test_data_template_views._resolve_project_scenario_path",
            return_value="/abs/scenario_1",
        ), patch(
            "pybirdai.utils.datapoint_test_run.excel_to_csv_converter.ExcelToCSVConverter.write_files",
            return_value=["/abs/scenario_1/thing.csv"],
        ):
            request = self.factory.post(
                "/api/test-data/import-excel/",
                {"file": self._upload(), "scenario_key": scenario.key},
            )
            response = import_excel_test_data(request)

        payload = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload["scenario"], "scenario_1")
        # Only file names are returned; server paths are not disclosed.
        self.assertEqual(payload["files"], ["thing.csv"])
        self.assertNotIn("/abs/", json.dumps(payload))

    def test_a_rejected_workbook_explains_why(self):
        with patch(
            "pybirdai.utils.datapoint_test_run.excel_to_csv_converter.ExcelToCSVConverter.convert_workbook",
            side_effect=UnsafeCellValueError("thing!A row 2 looks like a spreadsheet formula"),
        ):
            request = self.factory.post("/api/test-data/import-excel/", {"file": self._upload()})
            response = import_excel_test_data(request)

        self.assertEqual(response.status_code, 400)
        self.assertIn("looks like a spreadsheet formula", json.loads(response.content)["error"])

    def test_an_oversized_workbook_explains_the_limit(self):
        with patch(
            "pybirdai.utils.datapoint_test_run.excel_to_csv_converter.ExcelToCSVConverter.convert_workbook",
            side_effect=WorkbookTooLargeError("The workbook is larger than the 25 MB limit."),
        ):
            request = self.factory.post("/api/test-data/import-excel/", {"file": self._upload()})
            response = import_excel_test_data(request)

        self.assertEqual(response.status_code, 400)
        self.assertIn("25 MB limit", json.loads(response.content)["error"])
