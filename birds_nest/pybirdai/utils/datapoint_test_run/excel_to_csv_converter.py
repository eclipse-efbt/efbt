"""
Excel to CSV converter for BIRD test fixtures.

Turns a completed BIRD test data workbook back into the per-table CSV fixtures
the test runner loads.

The workbook and the fixtures deliberately differ in two ways, and reconciling
them is the whole job of this module:

- **Enumerated values.** The spreadsheet offers dropdowns reading
  ``"code: description"`` because a bare code is unreadable. The database and
  the fixture CSVs store only the code, so the description is stripped here.
- **Inherited columns.** A worksheet shows every field of a table including the
  ones inherited from its parent, which is useful context while editing. A
  fixture CSV is inserted straight into one table, so it carries that table's
  own columns only; inherited values belong to the parent table's own CSV.

Security
--------

The workbook arrives from outside the application and its output ends up in a
database, so this module treats it as untrusted input:

- **Size and shape are capped.** A spreadsheet is a zip archive, so a small
  upload can expand into a very large amount of data. The workbook size, the
  number of worksheets, and the number of rows are all limited.
- **Formula-looking cells are refused.** A value starting with ``=``, ``+`` or
  ``@`` is a spreadsheet formula, never valid BIRD data. Writing one into a CSV
  would run it if that CSV were later opened in a spreadsheet, so those cells
  are rejected by name instead of being written or silently rewritten.
- **Only worksheets that map to a real BIRD table are converted**, and the
  column set comes from the model, not from the header row, so an added column
  cannot introduce a new database column.
- **Output file names come from model class names**, never from the workbook.
- Loading the resulting CSVs uses parameterised inserts with column and table
  names taken from the Django model, so cell content cannot reach SQL as code.
"""

import csv
import datetime
import io
import logging
import os
from typing import Any, Dict, List, Optional, Tuple

from django.core.exceptions import SuspiciousFileOperation
from django.db import models
from django.utils._os import safe_join

from pybirdai.utils.datapoint_test_run.test_data_template_utils import (
    build_domain_code_lookup,
    extract_domain_dictionaries,
    get_bird_model_classes,
    get_local_field_export_map,
    get_model_by_table_name,
    resolve_domain_value_to_code,
)

logger = logging.getLogger(__name__)

#: Worksheets the template adds for guidance rather than data.
METADATA_SHEET_PREFIX = "_"

#: The worksheet holding the model name to worksheet title mapping.
TABLE_INDEX_SHEET = "_Table_Index"

#: Largest workbook accepted, in bytes. A spreadsheet is a compressed archive,
#: so this is a bound on the upload rather than on what it expands to.
MAX_WORKBOOK_BYTES = 25 * 1024 * 1024

#: Bounds on what a single workbook may expand into.
MAX_WORKSHEETS = 1000
MAX_ROWS_PER_SHEET = 20_000
MAX_TOTAL_ROWS = 100_000

#: Leading characters a spreadsheet treats as the start of a formula. A cell of
#: BIRD data never begins with one, so their presence means either a mistake or
#: an attempt at CSV injection.
FORMULA_PREFIXES = ("=", "+", "@")

#: Control characters that let a value break out of a CSV cell.
CONTROL_CHARACTERS = ("\t", "\r", "\n", "\x00")


class WorkbookTooLargeError(ValueError):
    """The workbook is larger, or holds more data, than the converter accepts."""


class UnsafeCellValueError(ValueError):
    """A cell holds a value that must not be written into a CSV file."""


class ExcelToCSVConverter:
    """Converts an edited BIRD test data workbook into fixture CSV files."""

    def __init__(self, allowed_root: Optional[str] = None):
        self.allowed_root = allowed_root
        self._models: Optional[Dict[str, Any]] = None
        self._total_rows = 0

    def _get_models(self) -> Dict[str, Any]:
        if self._models is None:
            self._models = get_bird_model_classes()
        return self._models

    def _check_workbook_size(self, source) -> None:
        """Refuse a workbook larger than the accepted upload size."""
        size = getattr(source, "size", None)
        if size is None:
            try:
                position = source.tell()
                source.seek(0, os.SEEK_END)
                size = source.tell()
                source.seek(position)
            except (AttributeError, OSError):
                try:
                    size = os.path.getsize(source)
                except (OSError, TypeError):
                    return

        if size is not None and size > MAX_WORKBOOK_BYTES:
            raise WorkbookTooLargeError(
                f"The workbook is larger than the {MAX_WORKBOOK_BYTES // (1024 * 1024)} MB limit."
            )

    def _load_workbook(self, source):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency is declared
            raise RuntimeError("openpyxl is required to read Excel test data") from exc

        self._check_workbook_size(source)

        # read_only streams the sheets instead of building them in memory, and
        # data_only keeps the cached result of any formula the user added, which
        # is the value they meant rather than the formula text.
        workbook = load_workbook(source, read_only=True, data_only=True)

        if len(workbook.sheetnames) > MAX_WORKSHEETS:
            workbook.close()
            raise WorkbookTooLargeError(
                f"The workbook has more than {MAX_WORKSHEETS} worksheets."
            )

        return workbook

    def _model_names_by_sheet(self, workbook) -> Dict[str, str]:
        """
        Map worksheet titles to BIRD model names.

        Worksheet titles are capped at 31 characters, so long model names are
        shortened when the template is written and the mapping is recorded on
        the _Table_Index sheet. Without that sheet we fall back to matching the
        title against the model names directly.
        """
        mapping = {}

        if TABLE_INDEX_SHEET in workbook.sheetnames:
            index_sheet = workbook[TABLE_INDEX_SHEET]
            for row in index_sheet.iter_rows(min_row=2, max_col=2, values_only=True):
                model_name, sheet_title = (row + (None, None))[:2]
                if model_name and sheet_title:
                    mapping[str(sheet_title)] = str(model_name)

        return mapping

    def _resolve_model(self, sheet_title: str, model_names_by_sheet: Dict[str, str]):
        model_name = model_names_by_sheet.get(sheet_title)
        if model_name:
            model = self._get_models().get(model_name)
            if model is not None:
                return model_name, model

        model = get_model_by_table_name(sheet_title)
        if model is not None:
            return model.__name__, model

        return None, None

    def _reject_unsafe_text(self, value: Any, location: str) -> None:
        """
        Refuse a cell whose text would be dangerous inside a CSV file.

        Rejecting rather than rewriting is deliberate: prefixing the value would
        change the data that ends up in the database, and a BIRD value never
        legitimately starts with a formula character. Negative numbers are not
        affected because a spreadsheet returns them as numbers, not text.
        """
        if not isinstance(value, str):
            return

        text = value.strip()
        if text.startswith(FORMULA_PREFIXES):
            raise UnsafeCellValueError(
                f"{location} looks like a spreadsheet formula ({text[:40]!r}). "
                "Remove the leading '=', '+' or '@' character."
            )

        if any(character in value for character in CONTROL_CHARACTERS):
            raise UnsafeCellValueError(
                f"{location} contains a line break or control character, which cannot be "
                "written to a CSV fixture."
            )

    def _convert_value(self, field, raw_value: Any, domain_lookup: Optional[Dict[str, str]]) -> Tuple[str, bool]:
        """
        Convert one cell into its CSV representation.

        Returns the text to write and whether a dropdown value was left
        unrecognised, so the caller can report it rather than silently writing
        something the fixture loader will reject.
        """
        if raw_value is None:
            return "", False

        if isinstance(raw_value, str):
            raw_value = raw_value.strip()
            if not raw_value or raw_value.upper() == "NULL":
                return "", False

        if domain_lookup is not None:
            code = resolve_domain_value_to_code(raw_value, domain_lookup)
            if code is not None:
                return code, False
            # Keep what the user wrote so the value is visible in the fixture
            # rather than silently dropped, but tell the caller about it.
            return str(raw_value), True

        if isinstance(raw_value, bool):
            return "True" if raw_value else "False", False
        if isinstance(raw_value, datetime.datetime):
            return raw_value.strftime("%Y-%m-%d %H:%M:%S"), False
        if isinstance(raw_value, datetime.date):
            return raw_value.strftime("%Y-%m-%d"), False

        if isinstance(raw_value, (int, float)):
            # Excel hands back every number as a float, so the field decides how
            # it should be written: a decimal keeps its point, and anything else
            # loses the ".0" that would otherwise appear on an identifier.
            if isinstance(field, (models.FloatField, models.DecimalField)):
                return str(float(raw_value)), False
            if float(raw_value).is_integer():
                return str(int(raw_value)), False

        return str(raw_value), False

    def convert_sheet(self, worksheet, model_class) -> Dict[str, Any]:
        """
        Convert one worksheet into the rows of one fixture CSV.

        Returns a dict with the CSV columns, the rows, and what was skipped.
        """
        rows_iterator = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iterator)
        except StopIteration:
            return {"columns": [], "rows": [], "ignored_columns": [], "unresolved_values": []}

        headers = [str(value).strip() if value is not None else "" for value in header_row]

        export_columns = get_local_field_export_map(model_class)
        domains = extract_domain_dictionaries(model_class)
        domain_lookups = {
            field_name: build_domain_code_lookup(domain) for field_name, domain in domains.items()
        }

        # Match each spreadsheet column to the table column it feeds, if any.
        column_for_header: Dict[int, Dict[str, Any]] = {}
        matched_columns = set()
        for header_index, header in enumerate(headers):
            if not header:
                continue
            for export_column in export_columns:
                if header in export_column["accepted_names"]:
                    column_for_header[header_index] = export_column
                    matched_columns.add(export_column["column"])
                    break

        ignored_columns = [
            header
            for header_index, header in enumerate(headers)
            if header and header_index not in column_for_header
        ]

        columns = [export_column["column"] for export_column in export_columns]
        unresolved_values: List[str] = []
        rows: List[List[str]] = []

        for row_number, row_values in enumerate(rows_iterator, start=2):
            if row_number > MAX_ROWS_PER_SHEET + 1:
                raise WorkbookTooLargeError(
                    f"Worksheet '{worksheet.title}' has more than {MAX_ROWS_PER_SHEET} data rows."
                )

            if row_values is None or all(
                value is None or (isinstance(value, str) and not value.strip()) for value in row_values
            ):
                continue

            values_by_column = {column: "" for column in columns}
            for header_index, export_column in column_for_header.items():
                if header_index >= len(row_values):
                    continue
                raw_value = row_values[header_index]
                location = f"{worksheet.title}!{headers[header_index]} row {row_number}"
                self._reject_unsafe_text(raw_value, location)

                field = export_column["field"]
                domain_lookup = domain_lookups.get(field.name)
                value, unresolved = self._convert_value(field, raw_value, domain_lookup)
                values_by_column[export_column["column"]] = value
                if unresolved:
                    unresolved_values.append(f"{location}: {raw_value!r}")

            self._total_rows += 1
            if self._total_rows > MAX_TOTAL_ROWS:
                raise WorkbookTooLargeError(
                    f"The workbook holds more than {MAX_TOTAL_ROWS} data rows in total."
                )

            rows.append([values_by_column[column] for column in columns])

        return {
            "columns": columns,
            "rows": rows,
            "ignored_columns": ignored_columns,
            "unresolved_values": unresolved_values,
        }

    def convert_workbook(self, source) -> Dict[str, Any]:
        """
        Convert a whole workbook into fixture CSV content.

        Args:
            source: A path or file-like object holding the .xlsx workbook

        Returns:
            Dict with:
            - files: {csv_filename: csv_text} for every table that had rows
            - tables: {model_name: row_count}
            - skipped_sheets: worksheets with no matching BIRD table
            - ignored_columns: {model_name: [header, ...]} not stored in that table
            - unresolved_values: dropdown values that could not be matched to a code
        """
        self._total_rows = 0
        workbook = self._load_workbook(source)
        try:
            model_names_by_sheet = self._model_names_by_sheet(workbook)

            files: Dict[str, str] = {}
            tables: Dict[str, int] = {}
            skipped_sheets: List[str] = []
            ignored_columns: Dict[str, List[str]] = {}
            unresolved_values: List[str] = []

            for sheet_title in workbook.sheetnames:
                if sheet_title.startswith(METADATA_SHEET_PREFIX):
                    continue

                model_name, model_class = self._resolve_model(sheet_title, model_names_by_sheet)
                if model_class is None:
                    skipped_sheets.append(sheet_title)
                    continue

                result = self.convert_sheet(workbook[sheet_title], model_class)
                if not result["rows"]:
                    continue

                buffer = io.StringIO()
                writer = csv.writer(buffer)
                writer.writerow(result["columns"])
                writer.writerows(result["rows"])

                files[f"{model_name.lower()}.csv"] = buffer.getvalue()
                tables[model_name] = len(result["rows"])
                if result["ignored_columns"]:
                    ignored_columns[model_name] = result["ignored_columns"]
                unresolved_values.extend(result["unresolved_values"])

            return {
                "files": files,
                "tables": tables,
                "skipped_sheets": skipped_sheets,
                "ignored_columns": ignored_columns,
                "unresolved_values": unresolved_values,
            }
        finally:
            workbook.close()

    def write_files(self, files: Dict[str, str], output_dir: str) -> List[str]:
        """Write converted CSV content into a scenario directory."""
        os.makedirs(output_dir, exist_ok=True)

        written = []
        for filename, content in files.items():
            if self.allowed_root:
                try:
                    csv_path = safe_join(output_dir, filename)
                except SuspiciousFileOperation as exc:
                    raise ValueError(f"Invalid fixture file name: {filename}") from exc
            else:
                csv_path = os.path.join(output_dir, filename)

            with open(csv_path, "w", encoding="utf-8", newline="") as f:
                f.write(content)
            written.append(csv_path)
            logger.info("Wrote fixture CSV %s", csv_path)

        return written


def convert_excel_to_csv_files(source) -> Dict[str, Any]:
    """Convenience wrapper around :class:`ExcelToCSVConverter`."""
    return ExcelToCSVConverter().convert_workbook(source)
