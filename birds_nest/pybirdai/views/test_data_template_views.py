"""
Test Data Template Views for generating Excel templates.

Provides an API endpoint to generate Excel workbooks with:
- One worksheet per BIRD data model table
- Column headers matching field names
- Dropdown validation for domain fields
"""

import io
import logging
import os
import re
from datetime import datetime

from django.conf import settings
from django.core.exceptions import SuspiciousFileOperation
from django.http import HttpResponse, JsonResponse
from django.utils._os import safe_join
from django.views.decorators.http import require_http_methods
from pybirdai.utils.datapoint_test_run.fixture_data_loader import (
    discover_fixture_scenarios,
    find_fixture_scenario,
)
from pybirdai.utils.secure_error_handling import SecureErrorHandler

logger = logging.getLogger(__name__)

EXCEL_MAX_SHEET_TITLE_LENGTH = 31
EXCEL_INVALID_SHEET_TITLE_CHARS = re.compile(r"[\\/*?:\[\]]")

#: Rows of dropdown validation to apply when a sheet has little or no data.
EXCEL_MIN_VALIDATED_ROWS = 1000

#: Default cap on how many existing rows are exported per table.
DEFAULT_MAX_DATA_ROWS = 1000


# Default tables to include in the template
DEFAULT_TEST_TABLES = [
    "PRTY",
    "FNNCL_CNTRCT",
    "CRDT_FCLTY",
    "INSTRMNT",
    "ENTTY_RL",
    "INSTRMNT_RL",
    "INSTRMNT_ENTTY_RL_ASSGNMNT",
    "CLLTRL",
    "CLLTRL_RL",
    "PRTCTN_ARRNGMNT",
    "PRTCTN_RCVD",
]


def _make_unique_excel_sheet_title(raw_title: str, used_titles: set[str]) -> str:
    """
    Return an Excel-safe, case-insensitively unique worksheet title.

    Excel limits worksheet titles to 31 characters. Letting openpyxl resolve
    collisions after truncation can produce a 32+ character title because it
    appends a numeric suffix to an already 31-character title.
    """
    base_title = EXCEL_INVALID_SHEET_TITLE_CHARS.sub("_", str(raw_title))
    base_title = base_title.strip().strip("'") or "Sheet"

    candidate = base_title[:EXCEL_MAX_SHEET_TITLE_LENGTH]
    suffix_number = 2
    while candidate.casefold() in used_titles:
        suffix = f"_{suffix_number}"
        prefix_length = EXCEL_MAX_SHEET_TITLE_LENGTH - len(suffix)
        candidate = f"{base_title[:prefix_length]}{suffix}"
        suffix_number += 1

    used_titles.add(candidate.casefold())
    return candidate


def _internal_json_error_response(exception: Exception, context: str, request, status: int = 500):
    """Hide implementation details from JSON error responses."""
    error_data = SecureErrorHandler.handle_exception(exception, context, request)
    return JsonResponse({"error": error_data["message"]}, status=status)


def _internal_http_error_response(exception: Exception, context: str, request, status: int = 500):
    """Hide implementation details from plain text responses."""
    error_data = SecureErrorHandler.handle_exception(exception, context, request)
    return HttpResponse(error_data["message"], status=status, content_type="text/plain")


def _resolve_project_scenario_path(scenario_path: str) -> str:
    """Keep fixture conversion requests inside the project tree."""
    try:
        return safe_join(str(settings.BASE_DIR), scenario_path)
    except SuspiciousFileOperation as exc:
        raise ValueError("Invalid fixture scenario path.") from exc


def _read_existing_rows(model_class, fields_metadata, domain_dicts, max_rows):
    """
    Read the rows currently stored for one table, formatted for the spreadsheet.

    Enumerated values are expanded from the stored code to the same
    "code: description" label the dropdown offers, so an exported cell matches
    its own validation list. Foreign keys are read as raw ids to avoid loading
    related objects.

    A table that cannot be read - because the model is not yet migrated, for
    instance - contributes no rows rather than failing the whole export.
    """
    from pybirdai.utils.datapoint_test_run.test_data_template_utils import (
        format_domain_value_for_cell,
        format_value_for_excel,
    )

    try:
        queryset = model_class.objects.all()[:max_rows]
        instances = list(queryset)
    except Exception as exc:
        logger.warning(
            "Skipping existing data for %s: %s",
            getattr(model_class, "__name__", model_class),
            exc,
        )
        return []

    value_attributes = []
    for field_meta in fields_metadata:
        # Reading the attname of a foreign key gives the stored id directly.
        value_attributes.append(
            f"{field_meta['name']}_id" if field_meta.get("is_foreign_key") else field_meta["name"]
        )

    rows = []
    for instance in instances:
        row = []
        for field_meta, attribute in zip(fields_metadata, value_attributes):
            value = getattr(instance, attribute, None)
            domain = domain_dicts.get(field_meta["name"])
            if domain and value is not None:
                row.append(format_domain_value_for_cell(value, domain))
            else:
                row.append(format_value_for_excel(value))
        rows.append(row)

    return rows


def _with_ancestor_models(selected_models, all_models):
    """
    Add the parent tables of any selected subtype to the selection.

    A worksheet carries only the columns its own table stores, so the fields a
    subtype inherits are entered on its parent's worksheet. Pulling the parents
    into the workbook is what keeps those fields reachable when someone asks for
    a subtype on its own; without them the inherited values would have nowhere
    to go.

    Returns the expanded selection and the names that were added.
    """
    name_by_model = {model: name for name, model in all_models.items()}
    expanded = dict(selected_models)
    added = []

    pending = list(selected_models.values())
    while pending:
        model_meta = getattr(pending.pop(), "_meta", None)
        for parent_model in getattr(model_meta, "parents", None) or {}:
            parent_name = name_by_model.get(parent_model)
            if parent_name is None or parent_name in expanded:
                continue
            expanded[parent_name] = parent_model
            added.append(parent_name)
            pending.append(parent_model)

    return expanded, added


@require_http_methods(["GET"])
def export_bird_excel_template(request):
    """
    Generate an Excel template for BIRD test data entry.

    Creates an Excel workbook with one worksheet per table, including:
    - Header row with the field names that table stores
    - The rows currently stored for that table, if any
    - Dropdown validation for domain fields (showing "code: description")
    - A reference sheet with all domain values
    - A table index mapping full model names to Excel-safe worksheet names

    A worksheet holds its own table's columns only - for a subtype, its link to
    the parent plus the fields declared on it - which is the same rule the
    fixture CSVs follow. The parent tables of any selected subtype are added to
    the workbook so that inherited fields can still be filled in, on the
    worksheet that actually stores them.

    Exported enumerated values are expanded from the stored code to the same
    "code: description" label the dropdown offers, so every cell satisfies its
    own validation. The upload endpoint reverses this.

    Query Parameters:
        tables: Comma-separated list of table names (e.g., "prty,instrmnt")
        include_all: If "true", include all BIRD tables (default: false)
        include_data: If "false", export an empty template (default: true)
        max_rows: Cap on exported rows per table (default: 1000)

    Returns:
        Excel file download
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter, quote_sheetname
        from openpyxl.workbook.defined_name import DefinedName
    except ImportError:
        return HttpResponse(
            "openpyxl is required for Excel export. Install with: pip install openpyxl",
            status=500,
            content_type="text/plain",
        )

    try:
        from pybirdai.utils.datapoint_test_run.test_data_template_utils import (
            get_bird_model_classes,
            extract_domain_dictionaries,
            get_model_fields_metadata,
            format_domain_value_for_dropdown,
        )

        # Parse query parameters
        tables_param = request.GET.get("tables", "")
        include_all = request.GET.get("include_all", "false").lower() == "true"
        include_data = request.GET.get("include_data", "true").lower() == "true"
        try:
            max_data_rows = max(0, int(request.GET.get("max_rows", DEFAULT_MAX_DATA_ROWS)))
        except ValueError:
            max_data_rows = DEFAULT_MAX_DATA_ROWS

        # Get model classes
        all_models = get_bird_model_classes()
        if not all_models:
            candidate_paths = [
                os.path.join(os.getcwd(), "pybirdai", "models", "bird_data_model.py"),
            ]
            try:
                from django.conf import settings

                candidate_paths.append(os.path.join(str(settings.BASE_DIR), "pybirdai", "models", "bird_data_model.py"))
            except Exception:
                pass

            # Preserve order while removing duplicates
            seen = set()
            deduped_candidates = []
            for path in candidate_paths:
                if path not in seen:
                    seen.add(path)
                    deduped_candidates.append(path)

            return JsonResponse(
                {
                    "error": (
                        "No BIRD data model tables were discovered. "
                        "Expected generated model file: pybirdai/models/bird_data_model.py"
                    ),
                    "hint": (
                        "Run the database setup/automode pipeline to generate bird_data_model.py, "
                        "then retry the template export."
                    ),
                    "checked_paths": deduped_candidates,
                },
                status=500,
            )

        if tables_param:
            # Use requested tables
            requested_tables = [t.strip().upper() for t in tables_param.split(",") if t.strip()]
            models_to_include = {name: model for name, model in all_models.items() if name.upper() in requested_tables}
            if not models_to_include:
                return JsonResponse(
                    {"error": f"No valid tables found. Available: {list(all_models.keys())[:10]}..."}, status=400
                )
        elif include_all:
            models_to_include = all_models
        else:
            # Use default tables
            models_to_include = {
                name: model
                for name, model in all_models.items()
                if name.upper() in [t.upper() for t in DEFAULT_TEST_TABLES]
            }

        # A subtype worksheet does not offer its inherited fields, so its parent
        # tables have to travel with it.
        models_to_include, parent_models_added = _with_ancestor_models(models_to_include, all_models)
        parent_only_models = set(parent_models_added)

        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        reserved_sheet_titles = {
            "_Instructions",
            "_Table_Index",
            "_Domains_Reference",
            "_Validation_Lists",
        }
        used_sheet_titles = {title.casefold() for title in reserved_sheet_titles}

        # Define styles
        header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        cell_font = Font(size=10)
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Track all domains for reference sheet
        all_domains_data = []
        table_sheet_index = []
        validation_ws = None
        validation_range_names = {}
        exported_row_count = 0

        # Create worksheet for each model
        for model_name in sorted(models_to_include.keys()):
            model_class = models_to_include[model_name]

            # Keep every title within Excel's 31-character limit, including
            # suffixes needed when long model names share the same prefix.
            ws_name = _make_unique_excel_sheet_title(
                model_name.lower(),
                used_sheet_titles,
            )
            ws = wb.create_sheet(title=ws_name)
            table_sheet_index.append((model_name, ws_name))

            # Get field metadata and domains
            fields_metadata = get_model_fields_metadata(model_class)
            domain_dicts = extract_domain_dictionaries(model_class)

            if not fields_metadata:
                continue

            # Read the rows first: the dropdown validation has to reach at least
            # as far down the sheet as the data does.
            data_rows = (
                _read_existing_rows(model_class, fields_metadata, domain_dicts, max_data_rows)
                if include_data and max_data_rows
                else []
            )
            exported_row_count += len(data_rows)
            validated_last_row = max(EXCEL_MIN_VALIDATED_ROWS, len(data_rows) + EXCEL_MIN_VALIDATED_ROWS)

            # Write header row
            for col_idx, field_meta in enumerate(fields_metadata, start=1):
                cell = ws.cell(row=1, column=col_idx, value=field_meta["name"])
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = header_alignment

                # Set column width
                col_letter = get_column_letter(col_idx)
                # Width based on header length, min 12, max 30
                width = min(30, max(12, len(field_meta["name"]) + 2))
                ws.column_dimensions[col_letter].width = width

                # Add dropdown validation for domain fields
                field_name = field_meta["name"]
                if field_name in domain_dicts:
                    domain = domain_dicts[field_name]

                    # Record domain for reference sheet
                    for code, description in domain.items():
                        all_domains_data.append(
                            {"table": model_name, "field": field_name, "code": code, "description": description}
                        )

                    sorted_domain_items = sorted(domain.items(), key=lambda x: str(x[0]))

                    # Create dropdown values in "code: description" format.
                    # Store them in a hidden sheet and refer to a workbook-level
                    # defined name. Inline validation lists are fragile because
                    # Excel limits the formula to 255 characters and treats
                    # commas and quotes in labels as syntax.
                    dropdown_values = []
                    for code, description in sorted_domain_items:
                        formatted = format_domain_value_for_dropdown(code, description, max_length=100)
                        dropdown_values.append(formatted)

                    if dropdown_values:
                        domain_key = tuple((str(code), str(description)) for code, description in sorted_domain_items)
                        range_name = validation_range_names.get(domain_key)
                        if range_name is None:
                            if validation_ws is None:
                                validation_ws = wb.create_sheet(title="_Validation_Lists")

                            validation_column = len(validation_range_names) + 1
                            validation_column_letter = get_column_letter(validation_column)
                            range_name = f"BIRD_Domain_{validation_column}"
                            validation_range_names[domain_key] = range_name

                            validation_ws.cell(
                                row=1,
                                column=validation_column,
                                value=range_name,
                            )
                            for row_idx, dropdown_value in enumerate(dropdown_values, start=2):
                                validation_ws.cell(
                                    row=row_idx,
                                    column=validation_column,
                                    value=dropdown_value,
                                )

                            range_reference = (
                                f"{quote_sheetname(validation_ws.title)}!"
                                f"${validation_column_letter}$2:"
                                f"${validation_column_letter}${len(dropdown_values) + 1}"
                            )
                            wb.defined_names.add(DefinedName(range_name, attr_text=range_reference))

                        dv = DataValidation(
                            type="list",
                            formula1=f"={range_name}",
                            showDropDown=False,  # False = show dropdown arrow
                            allow_blank=True,
                            showErrorMessage=True,
                            errorStyle="warning",
                            errorTitle="Invalid value",
                            error="Please select a value from the dropdown or enter a valid code.",
                        )
                        dv.add(f"{col_letter}2:{col_letter}{validated_last_row}")
                        ws.add_data_validation(dv)

            # Write the rows already in the database, below the header
            for row_offset, data_row in enumerate(data_rows, start=2):
                for col_idx, value in enumerate(data_row, start=1):
                    cell = ws.cell(row=row_offset, column=col_idx, value=value)
                    cell.font = cell_font
                    cell.border = thin_border

            # Freeze header row
            ws.freeze_panes = "A2"

            # Set row height for header
            ws.row_dimensions[1].height = 25

        # Create Domains Reference sheet
        if all_domains_data:
            ref_ws = wb.create_sheet(title="_Domains_Reference")

            # Headers
            ref_headers = ["Table", "Field", "Code", "Description"]
            for col_idx, header in enumerate(ref_headers, start=1):
                cell = ref_ws.cell(row=1, column=col_idx, value=header)
                cell.fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.border = thin_border

            # Set column widths
            ref_ws.column_dimensions["A"].width = 25
            ref_ws.column_dimensions["B"].width = 30
            ref_ws.column_dimensions["C"].width = 15
            ref_ws.column_dimensions["D"].width = 60

            # Data rows
            for row_idx, domain_item in enumerate(all_domains_data, start=2):
                ref_ws.cell(row=row_idx, column=1, value=domain_item["table"])
                ref_ws.cell(row=row_idx, column=2, value=domain_item["field"])
                ref_ws.cell(row=row_idx, column=3, value=domain_item["code"])
                # Clean up description
                desc = domain_item["description"].replace("_", " ")
                ref_ws.cell(row=row_idx, column=4, value=desc)

            # Freeze header
            ref_ws.freeze_panes = "A2"

        # Create a stable mapping from full BIRD model names to the necessarily
        # shortened worksheet titles.
        index_ws = wb.create_sheet(title="_Table_Index", index=0)
        index_headers = ["BIRD model/table", "Worksheet", "Included as"]
        for col_idx, header in enumerate(index_headers, start=1):
            cell = index_ws.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.border = thin_border

        for row_idx, (model_name, sheet_name) in enumerate(table_sheet_index, start=2):
            index_ws.cell(row=row_idx, column=1, value=model_name)
            index_ws.cell(row=row_idx, column=2, value=sheet_name)
            index_ws.cell(
                row=row_idx,
                column=3,
                value="parent table" if model_name in parent_only_models else "selected",
            )

        index_ws.column_dimensions["A"].width = 70
        index_ws.column_dimensions["B"].width = 35
        index_ws.column_dimensions["C"].width = 16
        index_ws.freeze_panes = "A2"
        index_ws.auto_filter.ref = f"A1:C{len(table_sheet_index) + 1}"

        if validation_ws is not None:
            # Users have the readable _Domains_Reference sheet; the raw range
            # source should not be edited or accidentally deleted.
            validation_ws.sheet_state = "veryHidden"

        # Create Instructions sheet
        instr_ws = wb.create_sheet(title="_Instructions", index=0)
        instructions = [
            ["BIRD Test Data Template"],
            [""],
            ["How to use this template:"],
            ["1. Each worksheet (tab) represents a BIRD data model table"],
            ["2. The header row contains the field names - do not modify"],
            ["3. Rows already in the database are filled in below the header - edit them or add your own"],
            ["4. Fields with dropdowns have enumerated values - select from the list"],
            ['5. Dropdown values read "CODE: description"; the description is for readability only'],
            ["6. Date fields should use ISO format: YYYY-MM-DD or YYYY-MM-DD HH:MM:SS"],
            ["7. Leave cells empty for NULL values"],
            ['8. Foreign key fields (starting with "the") should contain the unique ID of the referenced record'],
            ["9. See _Table_Index for full BIRD table names when worksheet names are shortened"],
            ["10. A worksheet holds the columns of its own table only. Where a table inherits from"],
            ["    a parent table, the inherited fields are entered on the parent's worksheet, and the"],
            ['    parent is listed in _Table_Index as "parent table".'],
            [""],
            ["To create CSV files for test fixtures:"],
            ["1. Complete your data entry in each table worksheet"],
            ["2. Save the workbook"],
            ['3. Upload it on the "Test Data Import" page'],
            ["4. The CSV fixtures are generated for you, with only the CODE stored for dropdown fields"],
            [""],
            ["The upload is the recommended route. Saving a worksheet as CSV by hand keeps the"],
            ["descriptions attached to dropdown values, which the fixture loader does not expect."],
            [""],
            ["See the _Domains_Reference sheet for a complete list of valid domain values."],
            [""],
            [f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'],
            [f"Tables included: {len(models_to_include)}"],
            [f"Existing rows included: {exported_row_count}"],
        ]

        if parent_models_added:
            instructions.append([""])
            instructions.append(
                [
                    "Parent tables added automatically so that inherited fields can be entered: "
                    + ", ".join(sorted(parent_models_added))
                ]
            )

        for row_idx, row_data in enumerate(instructions, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = instr_ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif value in ["How to use this template:", "To create CSV files for test fixtures:"]:
                    cell.font = Font(bold=True, size=11)

        instr_ws.column_dimensions["A"].width = 100

        # Generate response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"bird_test_data_template_{timestamp}.xlsx"

        response = HttpResponse(
            output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        logger.info(f"Generated Excel template with {len(models_to_include)} tables")
        return response

    except Exception as e:
        return _internal_http_error_response(e, "generating Excel template", request)


@require_http_methods(["GET"])
def list_available_tables(request):
    """
    List all available BIRD tables that can be included in the template.

    Returns:
        JSON response with list of table names
    """
    try:
        from pybirdai.utils.datapoint_test_run.test_data_template_utils import (
            get_bird_model_classes,
            DEFAULT_TEST_TABLES,
        )

        all_models = get_bird_model_classes()
        if not all_models:
            return JsonResponse(
                {
                    "error": (
                        "No BIRD data model tables were discovered. "
                        "Expected generated model file: pybirdai/models/bird_data_model.py"
                    ),
                    "tables": [],
                    "total": 0,
                    "default_tables": DEFAULT_TEST_TABLES,
                },
                status=500,
            )

        return JsonResponse(
            {
                "tables": sorted(all_models.keys()),
                "total": len(all_models),
                "default_tables": DEFAULT_TEST_TABLES,
            }
        )

    except Exception as e:
        return _internal_json_error_response(e, "listing test data template tables", request)


@require_http_methods(["GET"])
def test_data_import_page(request):
    """
    Render the page for uploading a completed test data workbook.

    The scenarios are passed in as choices rather than letting the page ask for
    a folder: the converted CSVs may only ever be written into a fixture
    scenario that already exists.
    """
    from django.shortcuts import render

    scenarios = discover_fixture_scenarios(str(settings.BASE_DIR))
    return render(
        request,
        "pybirdai/test_data_import.html",
        {
            "scenarios": [
                {
                    "key": scenario.key,
                    "label": (
                        f"{scenario.name} - {scenario.group}" if scenario.group else scenario.name
                    ),
                    "suite": scenario.suite,
                }
                for scenario in scenarios
            ]
        },
    )


@require_http_methods(["POST"])
def import_excel_test_data(request):
    """
    Convert an edited BIRD test data workbook into fixture CSV files.

    Dropdown cells read "code: description" but fixtures store the code alone,
    so the description is stripped here. Columns a worksheet inherits from a
    parent table are left out too - they belong in the parent's own CSV.

    The workbook is untrusted input whose output reaches a database, so it is
    size- and shape-limited and formula-looking cells are refused; see
    excel_to_csv_converter for the detail. Output goes either to the caller as a
    zip, or into one of the fixture scenarios that already exist - never to a
    path the request chooses.

    Request (multipart/form-data):
        file: the .xlsx workbook
        scenario_key: optional existing fixture scenario to write the CSVs into

    Returns:
        A zip of CSV files, or JSON describing what was written when
        scenario_key is given
    """
    uploaded_file = request.FILES.get("file")
    if uploaded_file is None:
        return JsonResponse({"error": "No workbook was uploaded."}, status=400)

    if not uploaded_file.name.lower().endswith((".xlsx", ".xlsm")):
        return JsonResponse({"error": "Please upload an .xlsx workbook."}, status=400)

    try:
        from pybirdai.utils.datapoint_test_run.excel_to_csv_converter import (
            ExcelToCSVConverter,
            UnsafeCellValueError,
            WorkbookTooLargeError,
        )

        # The target is resolved before any parsing happens, so an unusable
        # request is refused without reading the workbook at all.
        scenario_key = (request.POST.get("scenario_key") or "").strip()
        scenario = None
        if scenario_key:
            scenario = find_fixture_scenario(str(settings.BASE_DIR), scenario_key)
            if scenario is None:
                return JsonResponse({"error": "Fixture scenario was not found."}, status=404)

        converter = ExcelToCSVConverter(allowed_root=str(settings.BASE_DIR))
        try:
            result = converter.convert_workbook(uploaded_file)
        except (WorkbookTooLargeError, UnsafeCellValueError) as exc:
            logger.info("Rejected uploaded test data workbook: %s", exc)
            return JsonResponse(
                {"error": "The uploaded workbook is invalid or exceeds allowed limits."},
                status=400,
            )
        except Exception as exc:
            logger.info("Could not read uploaded test data workbook: %s", exc)
            return JsonResponse({"error": "The uploaded file could not be read as an Excel workbook."}, status=400)

        if not result["files"]:
            return JsonResponse(
                {
                    "error": "No test data rows were found in the workbook.",
                    "skipped_sheets": result["skipped_sheets"],
                },
                status=400,
            )

        if scenario is not None:
            try:
                resolved_path = _resolve_project_scenario_path(scenario.project_path)
            except ValueError:
                return JsonResponse({"error": "Invalid fixture scenario path."}, status=400)

            written = converter.write_files(result["files"], resolved_path)
            return JsonResponse(
                {
                    "success": True,
                    "files": [os.path.basename(path) for path in written],
                    "tables": result["tables"],
                    "row_count": sum(result["tables"].values()),
                    "skipped_sheets": result["skipped_sheets"],
                    "ignored_columns": result["ignored_columns"],
                    "unresolved_values": result["unresolved_values"],
                    "scenario": scenario.name,
                    "message": f"Wrote {len(written)} CSV fixture(s) to scenario '{scenario.name}'",
                }
            )

        import json
        import zipfile

        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
            for filename, content in sorted(result["files"].items()):
                archive.writestr(filename, content)
            archive.writestr(
                "_conversion_report.json",
                json.dumps(
                    {
                        "tables": result["tables"],
                        "skipped_sheets": result["skipped_sheets"],
                        "ignored_columns": result["ignored_columns"],
                        "unresolved_values": result["unresolved_values"],
                    },
                    indent=2,
                    sort_keys=True,
                ),
            )
        buffer.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        response = HttpResponse(buffer.read(), content_type="application/zip")
        response["Content-Disposition"] = f'attachment; filename="bird_test_data_csv_{timestamp}.zip"'
        logger.info(
            "Converted uploaded workbook into %s CSV fixture(s), %s row(s)",
            len(result["files"]),
            sum(result["tables"].values()),
        )
        return response

    except Exception as e:
        return _internal_json_error_response(e, "converting Excel test data to CSV", request)


@require_http_methods(["POST"])
def convert_sql_to_csv(request):
    """
    Convert an existing sql_inserts.sql file to CSV files.

    Request Body (JSON):
        scenario_path: Path to scenario directory containing sql_inserts.sql

    Returns:
        JSON response with conversion results
    """
    try:
        import json

        data = json.loads(request.body)
        scenario_path = data.get("scenario_path")

        if not scenario_path:
            return JsonResponse({"error": "scenario_path is required"}, status=400)
        try:
            scenario_path = _resolve_project_scenario_path(scenario_path)
        except ValueError:
            return JsonResponse({"error": "Invalid fixture scenario path."}, status=400)

        from pybirdai.utils.datapoint_test_run.sql_to_csv_converter import SQLToCSVConverter

        converter = SQLToCSVConverter(allowed_root=str(settings.BASE_DIR))
        output_files = converter.convert_scenario_in_place(scenario_path)

        return JsonResponse(
            {"success": True, "files": output_files, "message": f"Converted {len(output_files)} tables to CSV"}
        )

    except FileNotFoundError as e:
        logger.info("Scenario SQL fixture not found for CSV conversion: %s", e)
        return JsonResponse({"error": "Scenario SQL fixture file was not found."}, status=404)
    except Exception as e:
        SecureErrorHandler.handle_exception(e, "converting SQL fixtures to CSV", request)
        return JsonResponse({"error": "Failed to convert SQL fixtures to CSV."}, status=500)
