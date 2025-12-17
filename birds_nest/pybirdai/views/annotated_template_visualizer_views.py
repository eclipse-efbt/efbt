# coding=UTF-8
# Copyright (c) 2025 Bird Software Solutions Ltd
# This program and the accompanying materials
# are made available under the terms of the Eclipse Public License 2.0
# which accompanies this distribution, and is available at
# https://www.eclipse.org/legal/epl-2.0/
#
# SPDX-License-Identifier: EPL-2.0
#
# Contributors:
#    Benjamin Arfa - initial API and implementation
#

"""
Annotated Template Visualizer Module

Provides a read-only visualization of table axis ordinates with annotations.
Users can select a table and view its structure with variable/member annotations,
or export the annotated template to Excel.
"""

import logging
import io
from datetime import datetime

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.utils.html import escape

from pybirdai.models.bird_meta_data_model import (
    TABLE,
    AXIS,
    AXIS_ORDINATE,
    CELL_POSITION,
    TABLE_CELL,
    ORDINATE_ITEM,
    FRAMEWORK,
    FRAMEWORK_TABLE,
)
from django.db import models

logger = logging.getLogger(__name__)


def annotated_template_view(request):
    """
    Main view for the Annotated Template Visualizer.
    Displays dropdowns for Framework, Version, and Table selection.
    """
    # Get all frameworks for the dropdown
    frameworks = FRAMEWORK.objects.all().order_by('framework_id')

    context = {
        'frameworks': frameworks,
        'page_title': 'Annotated Template Visualizer',
    }

    return render(request, 'pybirdai/workflow/shared/annotated_template/index.html', context)


def annotated_template_embed_view(request, table_id):
    """
    Embed view for the Annotated Template Visualizer.
    Renders just the template visualization without selection form.
    Suitable for embedding in iframes (e.g., modals in step 7).
    """
    try:
        table = TABLE.objects.get(table_id=table_id)
    except TABLE.DoesNotExist:
        return render(request, 'pybirdai/workflow/shared/annotated_template/embed.html', {
            'error': f'Table not found: {table_id}',
            'table_id': table_id,
        })

    # Generate the HTML table directly
    table_html = generate_annotated_table_html(table_id)

    # Get summary info
    table_axes = AXIS.objects.filter(table_id=table)
    table_ordinates = AXIS_ORDINATE.objects.filter(axis_id__in=table_axes).select_related('axis_id')

    # Count rows, cols, z-axis
    row_count = 0
    col_count = 0
    z_ordinates = []

    for ordinate in table_ordinates:
        orientation = ordinate.axis_id.orientation if ordinate.axis_id else None
        if orientation in ['Y', '2']:
            row_count += 1
        elif orientation in ['X', '1']:
            col_count += 1
        elif orientation in ['Z', '3']:
            z_ordinates.append({
                'id': ordinate.axis_ordinate_id,
                'name': ordinate.name or ordinate.code or ordinate.axis_ordinate_id,
            })

    # Get cell count
    cell_positions = CELL_POSITION.objects.filter(axis_ordinate_id__in=table_ordinates)
    cell_count = cell_positions.values_list('cell_id', flat=True).distinct().count()

    context = {
        'table': table,
        'table_id': table_id,
        'table_html': table_html,
        'row_count': row_count,
        'col_count': col_count,
        'cell_count': cell_count,
        'z_ordinates': z_ordinates,
        'page_title': f'Annotated Template - {table.name}',
    }

    return render(request, 'pybirdai/workflow/shared/annotated_template/embed.html', context)


@require_http_methods(["GET"])
def get_annotated_template_api(request, table_id):
    """
    API endpoint that returns the annotated table structure as JSON.

    Returns:
    - table: Table metadata
    - row_ordinates: List of row ordinates with annotations
    - col_ordinates: List of column ordinates with annotations
    - z_ordinates: List of Z-axis ordinates (if any)
    - cell_matrix: Dict mapping (row_id, col_id) to cell info with annotations
    """
    try:
        table = TABLE.objects.get(table_id=table_id)
    except TABLE.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': f'Table not found: {table_id}'
        }, status=404)

    # Get axes for this table
    table_axes = AXIS.objects.filter(table_id=table)
    logger.info(f"[Annotated Template] Found {table_axes.count()} axes for table {table_id}")

    # Get all ordinates from those axes
    table_ordinates = AXIS_ORDINATE.objects.filter(
        axis_id__in=table_axes
    ).select_related('axis_id')

    # Get all ordinate items (variable/member annotations) for these ordinates
    ordinate_items = ORDINATE_ITEM.objects.filter(
        axis_ordinate_id__in=table_ordinates
    ).select_related('variable_id', 'member_id', 'axis_ordinate_id')

    # Build ordinate_id -> annotations mapping
    ordinate_annotations = {}
    for item in ordinate_items:
        ord_id = item.axis_ordinate_id.axis_ordinate_id if item.axis_ordinate_id else None
        if ord_id:
            if ord_id not in ordinate_annotations:
                ordinate_annotations[ord_id] = []
            annotation = {
                'variable_id': item.variable_id.variable_id if item.variable_id else None,
                'variable_code': item.variable_id.code if item.variable_id else None,
                'variable_name': item.variable_id.name if item.variable_id else None,
                'member_id': item.member_id.member_id if item.member_id else None,
                'member_code': item.member_id.code if item.member_id else None,
                'member_name': item.member_id.name if item.member_id else None,
            }
            ordinate_annotations[ord_id].append(annotation)

    # Get all cell positions for those ordinates
    cell_positions = CELL_POSITION.objects.filter(
        axis_ordinate_id__in=table_ordinates
    ).select_related('axis_ordinate_id', 'axis_ordinate_id__axis_id', 'cell_id')

    # Get unique cells
    cell_ids = cell_positions.values_list('cell_id', flat=True).distinct()
    table_cells = TABLE_CELL.objects.filter(cell_id__in=cell_ids)

    # Build cell_id -> positions mapping
    cell_to_positions = {}
    for pos in cell_positions:
        cell_id = pos.cell_id_id if hasattr(pos, 'cell_id_id') else (pos.cell_id.cell_id if pos.cell_id else None)
        if cell_id:
            if cell_id not in cell_to_positions:
                cell_to_positions[cell_id] = []
            cell_to_positions[cell_id].append(pos)

    # Build ordinates data structure
    ordinates_data = {}
    for ordinate in table_ordinates:
        orientation = ordinate.axis_id.orientation if ordinate.axis_id else 'Unknown'
        ordinates_data[ordinate.axis_ordinate_id] = {
            'id': ordinate.axis_ordinate_id,
            'name': ordinate.name or ordinate.code or ordinate.axis_ordinate_id,
            'code': ordinate.code,
            'axis_name': ordinate.axis_id.name if ordinate.axis_id else 'Unknown',
            'axis_orientation': orientation,
            'level': ordinate.level or 0,
            'order': ordinate.order or 0,
            'is_abstract': ordinate.is_abstract_header,
            'annotations': ordinate_annotations.get(ordinate.axis_ordinate_id, []),
        }

    # Build cell matrix
    cell_matrix = {}
    row_ordinates_set = set()
    col_ordinates_set = set()
    z_ordinates_set = set()

    for cell in table_cells:
        positions = cell_to_positions.get(cell.cell_id, [])

        row_ord_id = None
        col_ord_id = None
        z_ord_id = None

        for pos in positions:
            if pos.axis_ordinate_id and pos.axis_ordinate_id.axis_id:
                orientation = pos.axis_ordinate_id.axis_id.orientation
                ord_id = pos.axis_ordinate_id.axis_ordinate_id

                if orientation in ['Y', '2']:  # Row
                    row_ord_id = ord_id
                    row_ordinates_set.add(ord_id)
                elif orientation in ['X', '1']:  # Column
                    col_ord_id = ord_id
                    col_ordinates_set.add(ord_id)
                elif orientation in ['Z', '3']:  # Z-axis
                    z_ord_id = ord_id
                    z_ordinates_set.add(ord_id)

        if row_ord_id and col_ord_id:
            cell_key = f"{row_ord_id}|{col_ord_id}"
            cell_matrix[cell_key] = {
                'cell_id': cell.cell_id,
                'is_shaded': cell.is_shaded,
                'name': cell.name or '',
                'system_data_code': cell.system_data_code,
            }

    # Sort ordinates by level and order
    def sort_key(ord_id):
        data = ordinates_data.get(ord_id, {})
        return (data.get('level', 0), data.get('order', 0))

    row_ordinates = sorted(
        [ordinates_data[ord_id] for ord_id in row_ordinates_set if ord_id in ordinates_data],
        key=lambda x: (x.get('level', 0), x.get('order', 0))
    )
    col_ordinates = sorted(
        [ordinates_data[ord_id] for ord_id in col_ordinates_set if ord_id in ordinates_data],
        key=lambda x: (x.get('level', 0), x.get('order', 0))
    )
    z_ordinates = sorted(
        [ordinates_data[ord_id] for ord_id in z_ordinates_set if ord_id in ordinates_data],
        key=lambda x: (x.get('level', 0), x.get('order', 0))
    )

    return JsonResponse({
        'status': 'success',
        'table': {
            'table_id': table.table_id,
            'name': table.name,
            'code': table.code,
            'description': table.description,
            'version': table.version,
        },
        'row_ordinates': row_ordinates,
        'col_ordinates': col_ordinates,
        'z_ordinates': z_ordinates,
        'cell_matrix': cell_matrix,
        'total_cells': len(cell_matrix),
        'total_row_ordinates': len(row_ordinates),
        'total_col_ordinates': len(col_ordinates),
    })


@require_http_methods(["GET"])
def export_annotated_template_excel(request, table_id):
    """
    Export the annotated template as an Excel file matching the EBA annotated template format.
    Uses openpyxl to create a formatted Excel workbook.

    Format matches resources/annotated_templates/:
    - Row 1: Title
    - Row 2: Z-axis variable (if applicable)
    - Row 3: Z-axis member (if applicable)
    - Row 5: "Columns" label
    - Rows 6-7: Column header names
    - Row 8: Column codes
    - Row 9+: Data rows with cell IDs and dimension columns at end
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            "openpyxl is required for Excel export. Install with: pip install openpyxl",
            status=500,
            content_type='text/plain'
        )

    try:
        table = TABLE.objects.get(table_id=table_id)
    except TABLE.DoesNotExist:
        return HttpResponse(f'Table not found: {table_id}', status=404)

    # Get axes and ordinates
    table_axes = AXIS.objects.filter(table_id=table)
    table_ordinates = AXIS_ORDINATE.objects.filter(
        axis_id__in=table_axes
    ).select_related('axis_id')

    # Get ordinate items for annotations
    ordinate_items = ORDINATE_ITEM.objects.filter(
        axis_ordinate_id__in=table_ordinates
    ).select_related('variable_id', 'member_id', 'axis_ordinate_id', 'variable_id__domain_id')

    # Build ordinate_id -> ordinate items list (full item data)
    ordinate_items_map = {}
    for item in ordinate_items:
        ord_id = item.axis_ordinate_id.axis_ordinate_id if item.axis_ordinate_id else None
        if ord_id:
            if ord_id not in ordinate_items_map:
                ordinate_items_map[ord_id] = []
            # Store full item info for dimension columns
            var_code = item.variable_id.code if item.variable_id else ''
            var_name = item.variable_id.name if item.variable_id else ''
            mem_code = item.member_id.code if item.member_id else ''
            mem_name = item.member_id.name if item.member_id else ''
            domain_code = item.variable_id.domain_id.code if item.variable_id and item.variable_id.domain_id else ''

            ordinate_items_map[ord_id].append({
                'variable_code': var_code,
                'variable_name': var_name,
                'member_code': mem_code,
                'member_name': mem_name,
                'domain_code': domain_code,
            })

    # Get cell positions
    cell_positions = CELL_POSITION.objects.filter(
        axis_ordinate_id__in=table_ordinates
    ).select_related('axis_ordinate_id', 'axis_ordinate_id__axis_id', 'cell_id')

    cell_ids = cell_positions.values_list('cell_id', flat=True).distinct()
    table_cells = TABLE_CELL.objects.filter(cell_id__in=cell_ids)

    # Build cell_id -> positions mapping
    cell_to_positions = {}
    for pos in cell_positions:
        cell_id = pos.cell_id_id if hasattr(pos, 'cell_id_id') else (pos.cell_id.cell_id if pos.cell_id else None)
        if cell_id:
            if cell_id not in cell_to_positions:
                cell_to_positions[cell_id] = []
            cell_to_positions[cell_id].append(pos)

    # Build ordinates data and separate by axis
    ordinates_data = {}
    z_ordinates = []
    row_ordinates_list = []
    col_ordinates_list = []

    for ordinate in table_ordinates:
        orientation = ordinate.axis_id.orientation if ordinate.axis_id else 'Unknown'
        path = ordinate.path or ''
        path_level = path.count('.') if path else 0
        # Extract ordinate code from ID (e.g., "..._X_0220" -> "0220")
        ord_code = ordinate.code or ''
        if not ord_code and '_' in ordinate.axis_ordinate_id:
            parts = ordinate.axis_ordinate_id.split('_')
            ord_code = parts[-1] if parts else ''

        ord_data = {
            'id': ordinate.axis_ordinate_id,
            'name': ordinate.name or ordinate.code or ordinate.axis_ordinate_id,
            'code': ord_code,
            'orientation': orientation,
            'level': path_level,
            'order': ordinate.order or 0,
            'path': path,
        }
        ordinates_data[ordinate.axis_ordinate_id] = ord_data

        if orientation == 'Z':
            z_ordinates.append(ordinate.axis_ordinate_id)
        elif orientation in ['Y', '2']:
            row_ordinates_list.append(ordinate.axis_ordinate_id)
        elif orientation in ['X', '1']:
            col_ordinates_list.append(ordinate.axis_ordinate_id)

    # Build cell matrix
    cell_matrix = {}
    row_ordinates_set = set()
    col_ordinates_set = set()

    for cell in table_cells:
        positions = cell_to_positions.get(cell.cell_id, [])
        row_ord_id = None
        col_ord_id = None

        for pos in positions:
            if pos.axis_ordinate_id and pos.axis_ordinate_id.axis_id:
                orientation = pos.axis_ordinate_id.axis_id.orientation
                ord_id = pos.axis_ordinate_id.axis_ordinate_id

                if orientation in ['Y', '2']:
                    row_ord_id = ord_id
                    row_ordinates_set.add(ord_id)
                elif orientation in ['X', '1']:
                    col_ord_id = ord_id
                    col_ordinates_set.add(ord_id)

        if row_ord_id and col_ord_id:
            cell_matrix[(row_ord_id, col_ord_id)] = {
                'cell_id': cell.cell_id,
                'is_shaded': cell.is_shaded,
                'name': cell.name or '',
            }

    # Sort ordinates
    row_ordinates = sorted(
        [ord_id for ord_id in row_ordinates_set if ord_id in ordinates_data],
        key=lambda x: (ordinates_data[x].get('level', 0), ordinates_data[x].get('order', 0))
    )
    col_ordinates = sorted(
        [ord_id for ord_id in col_ordinates_set if ord_id in ordinates_data],
        key=lambda x: (ordinates_data[x].get('level', 0), ordinates_data[x].get('order', 0))
    )

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table.code or 'Annotated Template'

    # Styles
    header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    row_header_fill = PatternFill(start_color='334155', end_color='334155', fill_type='solid')
    corner_fill = PatternFill(start_color='4C51BF', end_color='4C51BF', fill_type='solid')
    shaded_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    dim_header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True)
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Row 1: Title
    ws['A1'] = f"{table.code or table_id} - {table.name or 'Annotated Template'}"
    ws['A1'].font = Font(size=14, bold=True)

    # Row 2+: Z-axis info (display ALL Z ordinate items)
    z_row = 2
    if z_ordinates:
        for z_ord_id in z_ordinates:
            z_items = ordinate_items_map.get(z_ord_id, [])
            for z_item in z_items:
                # Variable info: variable_name (variable_code)
                var_name = z_item['variable_name'] or z_item['variable_code'] or ''
                var_code = z_item['variable_code'] or ''
                z_var_info = f"{var_name} ({var_code})" if var_code else var_name
                ws.cell(row=z_row, column=4, value=z_var_info)
                ws.cell(row=z_row, column=4).font = bold_font
                z_row += 1
                # Member info: member_name (member_code)
                mem_name = z_item['member_name'] or z_item['member_code'] or ''
                mem_code = z_item['member_code'] or ''
                z_mem_info = f"{mem_name} ({mem_code})" if mem_code else mem_name
                ws.cell(row=z_row, column=4, value=z_mem_info)
                z_row += 1

    # Dynamic row positioning based on Z items
    columns_label_row = max(z_row + 1, 5)
    ws.cell(row=columns_label_row, column=4, value='Columns')
    ws.cell(row=columns_label_row, column=4).font = bold_font

    # Column header names
    header_row = columns_label_row + 1
    for col_idx, col_ord_id in enumerate(col_ordinates, start=4):
        col_data = ordinates_data.get(col_ord_id, {})
        header_text = col_data.get('name', col_ord_id)
        cell = ws.cell(row=header_row, column=col_idx, value=header_text)
        cell.fill = header_fill
        cell.font = white_font
        cell.border = thin_border
        cell.alignment = center_alignment
        # Set column width based on header length (min 20, max 40)
        col_width = min(max(len(header_text) + 4, 20), 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Column codes
    code_row = header_row + 1
    for col_idx, col_ord_id in enumerate(col_ordinates, start=4):
        col_data = ordinates_data.get(col_ord_id, {})
        cell = ws.cell(row=code_row, column=col_idx, value=col_data.get('code', ''))
        cell.border = thin_border
        cell.alignment = center_alignment

    # Collect Z-axis variables first (to exclude from X and Y axis - they're shown at top)
    z_dim_vars = set()
    for z_ord_id in z_ordinates:
        for item in ordinate_items_map.get(z_ord_id, []):
            if item['variable_code']:
                z_dim_vars.add(item['variable_code'])

    # Collect dimension variables SEPARATELY by axis
    # X-axis (column) dimension variables - exclude Z-axis variables
    x_dim_vars = set()
    for col_ord_id in col_ordinates:
        for item in ordinate_items_map.get(col_ord_id, []):
            if item['variable_code'] and item['variable_code'] not in z_dim_vars:
                x_dim_vars.add(item['variable_code'])
    x_dim_vars_list = sorted(x_dim_vars)

    # Y-axis (row) dimension variables - exclude Z-axis variables
    y_dim_vars = set()
    for row_ord_id in row_ordinates:
        for item in ordinate_items_map.get(row_ord_id, []):
            if item['variable_code'] and item['variable_code'] not in z_dim_vars:
                y_dim_vars.add(item['variable_code'])
    y_dim_vars_list = sorted(y_dim_vars)

    # Build Y-axis variable code -> variable name lookup
    y_var_names = {}
    for row_ord_id in row_ordinates:
        for item in ordinate_items_map.get(row_ord_id, []):
            if item['variable_code'] and item['variable_code'] not in y_var_names:
                y_var_names[item['variable_code']] = item['variable_name'] or item['variable_code']

    # Y-axis dimension column headers (at the end, after data columns)
    dim_start_col = 4 + len(col_ordinates) + 2  # Leave a gap
    for dim_idx, dim_var in enumerate(y_dim_vars_list):
        # Header: variable_name (variable_code)
        var_name = y_var_names.get(dim_var, dim_var)
        header_text = f"{var_name} ({dim_var})" if var_name != dim_var else dim_var
        cell = ws.cell(row=header_row, column=dim_start_col + dim_idx, value=header_text)
        cell.fill = dim_header_fill
        cell.font = white_font
        cell.border = thin_border
        cell.alignment = center_alignment
        # Set width based on header length (min 35, max 50)
        dim_col_width = min(max(len(header_text) + 4, 35), 50)
        ws.column_dimensions[get_column_letter(dim_start_col + dim_idx)].width = dim_col_width

    # Set column widths for first 3 columns
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 45  # Row names
    ws.column_dimensions['C'].width = 45  # Also row names (X-axis labels below)

    # Set row heights for better readability
    ws.row_dimensions[header_row].height = 40  # Column headers
    ws.row_dimensions[code_row].height = 20    # Column codes

    # Freeze panes so headers stay visible when scrolling
    ws.freeze_panes = ws.cell(row=code_row + 1, column=4)

    # Data rows (dynamic start based on header rows)
    data_start_row = code_row + 1
    for row_idx, row_ord_id in enumerate(row_ordinates, start=data_start_row):
        row_data = ordinates_data.get(row_ord_id, {})

        # Column A: "Rows" label (only on first row)
        if row_idx == data_start_row:
            ws.cell(row=row_idx, column=1, value='Rows')
            ws.cell(row=row_idx, column=1).font = bold_font

        # Column B: Row name with hierarchy indent
        level = row_data.get('level', 0)
        indent = '  ' * level
        header_text = row_data.get('name', row_ord_id)
        cell = ws.cell(row=row_idx, column=2, value=f"{indent}{header_text}")
        cell.fill = row_header_fill
        cell.font = white_font
        cell.border = thin_border
        cell.alignment = left_alignment

        # Column C: Row code
        cell = ws.cell(row=row_idx, column=3, value=row_data.get('code', ''))
        cell.border = thin_border
        cell.alignment = center_alignment

        # Data cells (columns D onwards)
        for col_idx, col_ord_id in enumerate(col_ordinates, start=4):
            cell_key = (row_ord_id, col_ord_id)
            cell_info = cell_matrix.get(cell_key)

            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = center_alignment

            if cell_info:
                # Format: cell_id + newline + metric symbol
                cell_id = cell_info.get('cell_id', '')
                # Extract numeric part from cell_id if it starts with REF_
                if cell_id.startswith('REF_'):
                    display_id = cell_id.replace('REF_', '')
                else:
                    display_id = cell_id
                cell.value = f"{display_id}\n$"
                if cell_info.get('is_shaded'):
                    cell.fill = shaded_fill
            else:
                cell.value = ''

        # Y-axis dimension columns at end of row - ONLY row ordinate items
        row_items = ordinate_items_map.get(row_ord_id, [])
        row_dim_values = {item['variable_code']: item for item in row_items if item['variable_code']}

        for dim_idx, dim_var in enumerate(y_dim_vars_list):
            if dim_var in row_dim_values:
                item = row_dim_values[dim_var]
                # Format: member_name (member_code)
                mem_name = item['member_name'] or item['member_code'] or '*'
                mem_code = item['member_code'] or ''
                dim_text = f"{mem_name} ({mem_code})" if mem_code else mem_name
                cell = ws.cell(row=row_idx, column=dim_start_col + dim_idx, value=dim_text)
                cell.border = thin_border
                cell.alignment = left_alignment

    # X-axis ordinate item rows (BELOW the data rows, in front of X-axis columns)
    # One row per unique X-axis variable
    x_dim_start_row = data_start_row + len(row_ordinates) + 1  # After data rows + gap

    # Build variable code -> variable name lookup
    x_var_names = {}
    for col_ord_id in col_ordinates:
        for item in ordinate_items_map.get(col_ord_id, []):
            if item['variable_code'] and item['variable_code'] not in x_var_names:
                x_var_names[item['variable_code']] = item['variable_name'] or item['variable_code']

    for x_dim_idx, x_dim_var in enumerate(x_dim_vars_list):
        current_row = x_dim_start_row + x_dim_idx
        # Label in column B (wider): variable_name (variable_code)
        var_name = x_var_names.get(x_dim_var, x_dim_var)
        label_text = f"{var_name} ({x_dim_var})" if var_name != x_dim_var else x_dim_var
        label_cell = ws.cell(row=current_row, column=2, value=label_text)
        label_cell.font = white_font
        label_cell.fill = dim_header_fill
        label_cell.border = thin_border
        label_cell.alignment = left_alignment

        # For each column, show that column's ordinate item for this variable
        for col_idx, col_ord_id in enumerate(col_ordinates, start=4):
            col_items = ordinate_items_map.get(col_ord_id, [])
            # Find item with this variable
            item_for_var = None
            for item in col_items:
                if item['variable_code'] == x_dim_var:
                    item_for_var = item
                    break

            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            cell.alignment = left_alignment
            if item_for_var:
                # Format: member_name (member_code)
                mem_name = item_for_var['member_name'] or item_for_var['member_code'] or '*'
                mem_code = item_for_var['member_code'] or ''
                cell.value = f"{mem_name} ({mem_code})" if mem_code else mem_name

    # Create response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename
    safe_table_code = (table.code or table_id).replace('/', '_').replace('\\', '_').replace('.', '_')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"annotated_template_{safe_table_code}_{timestamp}.xlsx"

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


def generate_annotated_table_html(table_id):
    """
    Generate HTML representation of the annotated table.
    Returns HTML string that can be embedded in a template.
    """
    try:
        table = TABLE.objects.get(table_id=table_id)
    except TABLE.DoesNotExist:
        return '<div class="alert alert-danger">Table not found</div>'

    # Get axes and ordinates
    table_axes = AXIS.objects.filter(table_id=table)
    table_ordinates = AXIS_ORDINATE.objects.filter(
        axis_id__in=table_axes
    ).select_related('axis_id')

    # Get ordinate items for annotations
    ordinate_items = ORDINATE_ITEM.objects.filter(
        axis_ordinate_id__in=table_ordinates
    ).select_related('variable_id', 'member_id', 'axis_ordinate_id')

    # Build ordinate_id -> annotations mapping
    ordinate_annotations = {}
    for item in ordinate_items:
        ord_id = item.axis_ordinate_id.axis_ordinate_id if item.axis_ordinate_id else None
        if ord_id:
            if ord_id not in ordinate_annotations:
                ordinate_annotations[ord_id] = []
            var_code = item.variable_id.code if item.variable_id else ''
            mem_code = item.member_id.code if item.member_id else ''
            var_name = item.variable_id.name if item.variable_id else ''
            mem_name = item.member_id.name if item.member_id else ''

            # Use name for display if available, fallback to code
            # Truncate long values for cleaner display
            def truncate(text, max_len=25):
                if text and len(text) > max_len:
                    return text[:max_len-3] + '...'
                return text or ''

            # Prefer name for display, fallback to code
            var_display = truncate(var_name) if var_name else truncate(var_code)
            mem_display = truncate(mem_name) if mem_name else truncate(mem_code)

            # Build display string
            if var_display and mem_display:
                display_text = f"{var_display}:{mem_display}"
            elif var_display:
                display_text = var_display
            else:
                display_text = mem_display

            if var_code or mem_code:
                ordinate_annotations[ord_id].append({
                    'var_code': var_code,
                    'mem_code': mem_code,
                    'var_name': var_name,
                    'mem_name': mem_name,
                    'display': display_text,
                    'tooltip': f"{var_name or var_code}: {mem_name or mem_code}",
                })

    # Get cell positions
    cell_positions = CELL_POSITION.objects.filter(
        axis_ordinate_id__in=table_ordinates
    ).select_related('axis_ordinate_id', 'axis_ordinate_id__axis_id', 'cell_id')

    cell_ids = cell_positions.values_list('cell_id', flat=True).distinct()
    table_cells = TABLE_CELL.objects.filter(cell_id__in=cell_ids)

    # Build cell_id -> positions mapping
    cell_to_positions = {}
    for pos in cell_positions:
        cell_id = pos.cell_id_id if hasattr(pos, 'cell_id_id') else (pos.cell_id.cell_id if pos.cell_id else None)
        if cell_id:
            if cell_id not in cell_to_positions:
                cell_to_positions[cell_id] = []
            cell_to_positions[cell_id].append(pos)

    # Build ordinates data
    ordinates_data = {}
    for ordinate in table_ordinates:
        orientation = ordinate.axis_id.orientation if ordinate.axis_id else 'Unknown'
        ordinates_data[ordinate.axis_ordinate_id] = {
            'id': ordinate.axis_ordinate_id,
            'name': ordinate.name or ordinate.code or ordinate.axis_ordinate_id,
            'orientation': orientation,
            'level': ordinate.level or 0,
            'order': ordinate.order or 0,
            'annotations': ordinate_annotations.get(ordinate.axis_ordinate_id, []),
        }

    # Build cell matrix
    cell_matrix = {}
    row_ordinates_set = set()
    col_ordinates_set = set()

    for cell in table_cells:
        positions = cell_to_positions.get(cell.cell_id, [])

        row_ord_id = None
        col_ord_id = None

        for pos in positions:
            if pos.axis_ordinate_id and pos.axis_ordinate_id.axis_id:
                orientation = pos.axis_ordinate_id.axis_id.orientation
                ord_id = pos.axis_ordinate_id.axis_ordinate_id

                if orientation in ['Y', '2']:
                    row_ord_id = ord_id
                    row_ordinates_set.add(ord_id)
                elif orientation in ['X', '1']:
                    col_ord_id = ord_id
                    col_ordinates_set.add(ord_id)

        if row_ord_id and col_ord_id:
            cell_matrix[(row_ord_id, col_ord_id)] = {
                'cell_id': cell.cell_id,
                'is_shaded': cell.is_shaded,
                'name': cell.name or '',
            }

    # Sort ordinates
    row_ordinates = sorted(
        [ordinates_data[ord_id] for ord_id in row_ordinates_set if ord_id in ordinates_data],
        key=lambda x: (x.get('level', 0), x.get('order', 0))
    )
    col_ordinates = sorted(
        [ordinates_data[ord_id] for ord_id in col_ordinates_set if ord_id in ordinates_data],
        key=lambda x: (x.get('level', 0), x.get('order', 0))
    )

    # Generate HTML
    html = ['<table class="annotated-table table-bordered">']

    # Header row
    html.append('<thead><tr>')
    html.append('<th class="corner-cell">Row / Column</th>')

    for col_ord in col_ordinates:
        name = escape(col_ord.get('name', 'N/A'))
        annotations = col_ord.get('annotations', [])
        annotation_html = ''
        if annotations:
            annotation_parts = [escape(a['display']) for a in annotations]
            annotation_html = f'<div class="ordinate-annotation">{" | ".join(annotation_parts)}</div>'
        tooltip_parts = [escape(a['tooltip']) for a in annotations] if annotations else []
        tooltip = escape(f"{name}\n" + "\n".join(tooltip_parts) if tooltip_parts else name)

        html.append(
            f'<th class="col-header" title="{tooltip}">'
            f'<div class="ordinate-name">{name}</div>'
            f'{annotation_html}</th>'
        )

    html.append('</tr></thead>')

    # Body rows
    html.append('<tbody>')

    for row_ord in row_ordinates:
        html.append('<tr>')

        name = escape(row_ord.get('name', 'N/A'))
        level = row_ord.get('level', 0)
        annotations = row_ord.get('annotations', [])
        annotation_html = ''
        if annotations:
            annotation_parts = [escape(a['display']) for a in annotations]
            annotation_html = f'<div class="ordinate-annotation">{" | ".join(annotation_parts)}</div>'
        tooltip_parts = [escape(a['tooltip']) for a in annotations] if annotations else []
        tooltip = escape(f"{name}\n" + "\n".join(tooltip_parts) if tooltip_parts else name)

        indent_style = f'padding-left: {level * 20 + 8}px;' if level > 0 else ''

        html.append(
            f'<th class="row-header" style="{indent_style}" title="{tooltip}">'
            f'<div class="ordinate-name">{name}</div>'
            f'{annotation_html}</th>'
        )

        # Data cells
        for col_ord in col_ordinates:
            cell_key = (row_ord['id'], col_ord['id'])
            cell_info = cell_matrix.get(cell_key)

            if cell_info:
                is_shaded = cell_info.get('is_shaded')
                cell_name = escape(cell_info.get('name', ''))
                shade_class = 'cell-shaded' if is_shaded else ''
                cell_tooltip = escape(f"Cell: {cell_name}" if cell_name else "Data cell")

                html.append(
                    f'<td class="data-cell {shade_class}" title="{cell_tooltip}">'
                    f'<span class="cell-indicator">X</span></td>'
                )
            else:
                html.append(
                    '<td class="data-cell cell-empty" title="No data in this cell">'
                    '<span class="empty-indicator">-</span></td>'
                )

        html.append('</tr>')

    html.append('</tbody>')
    html.append('</table>')

    return '\n'.join(html)
